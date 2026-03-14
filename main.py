from fastapi import FastAPI, Request, Form, Cookie, Query, BackgroundTasks, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, Response, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List
import uuid, time, json, os, re, bcrypt, logging, collections, hashlib, secrets, hmac
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

from database import (init_db, get_conn, get_influencers, get_influencer, get_influencer_posts, get_influencer_reels,
                      get_stats, get_public_stats, get_public_influencers,
                      get_manual, save_manual, get_advertisers, get_refresh_status,
                      update_influencer_stats, get_advertiser_by_username,
                      add_advertiser as db_add_advertiser, delete_advertiser as db_delete_advertiser,
                      update_advertiser_plan,
                      get_hashtags, get_collect_jobs, add_hashtag as db_add_hashtag,
                      delete_hashtag as db_delete_hashtag, update_hashtag_status,
                      add_collect_job, update_collect_job,
                      get_influencer_by_username, update_influencer_profile,
                      get_favorites, get_favorite_pks, toggle_favorite,
                      get_campaigns, create_campaign, get_campaign, get_campaign_influencers,
                      add_to_campaign, remove_from_campaign, delete_campaign,
                      add_cron_log, get_cron_logs, get_auto_hashtags,
                      get_url_stats, batch_upsert_from_excel)

# ── 해시태그 활동 유형 분석 헬퍼 ──
_SELLER_KW = {
    "공동구매","공구","단독공구","공구오픈","공구마감","공구진행","공구알림",
    "마감임박","품절임박","재오픈","재입고","선착순","추가생산","예약구매","한정수량",
    "오픈예정","마감","품절","솔드아웃","추가주문",
    "스마트스토어","쇼핑몰","내쇼핑몰","구매대행","셀러",
    "링크인바이오","프로필링크","네이버쇼핑","카카오쇼핑",
    "라이브커머스","네이버라이브","카카오라이브","라이브쇼핑","쇼핑라이브",
}
_AD_KW = {
    "협찬","체험단","광고","유료광고","제공","제공받음","후원","지원받음",
    "ad","sponsored","partnership","콜라보","콜라보레이션",
    "ppl","브랜드협찬","브랜드체험","무료체험","서포터즈","앰배서더",
    "솔직후기","내돈내산","리뷰","사용후기","실사용","내돈내구매",
}

def _analyze_activity(htags: list) -> dict:
    """해시태그 사용 빈도 가중합으로 공구셀러 vs 광고수신 vs 순수콘텐츠 비중 계산"""
    def _match(tag: str, kw_set) -> bool:
        t = tag.lower().replace(" ", "").replace("#", "")
        return any(kw in t for kw in kw_set)

    seller_score = sum(h["count"] for h in htags if _match(h["tag"], _SELLER_KW))
    ad_score     = sum(h["count"] for h in htags if _match(h["tag"], _AD_KW))
    total_tagged = seller_score + ad_score
    content_score = max(0, sum(h["count"] for h in htags) - total_tagged)
    grand = seller_score + ad_score + content_score or 1

    return {
        "seller_pct":  round(seller_score  / grand * 100),
        "ad_pct":      round(ad_score      / grand * 100),
        "content_pct": round(content_score / grand * 100),
        "seller_tags": [h["tag"] for h in htags if _match(h["tag"], _SELLER_KW)][:8],
        "ad_tags":     [h["tag"] for h in htags if _match(h["tag"], _AD_KW)][:8],
        "primary": (
            "seller"  if seller_score > ad_score * 1.5 and seller_score > 0
            else "ad" if ad_score > seller_score * 1.5 and ad_score > 0
            else "mixed"  if total_tagged > 0
            else "content"
        ),
        "has_data": len(htags) > 0,
    }


app = FastAPI()

# ─── CORS: 크롬 확장 프로그램만 허용, 와일드카드(*) 제거 ───
_CORS_ORIGINS = [o.strip() for o in os.getenv("CORS_ORIGINS", "").split(",") if o.strip()]
app.add_middleware(CORSMiddleware,
    allow_origins=_CORS_ORIGINS or [],          # 기본: 외부 origin 차단 (same-origin만 허용)
    allow_origin_regex=r"^chrome-extension://.*$",  # 크롬 확장은 패턴으로 허용
    allow_methods=["GET","POST"],
    allow_headers=["X-Api-Key", "Content-Type"],
    allow_credentials=False,                    # credentials 불필요 시 false
)

# ─── 봇 차단 / Rate Limiting ───────────────────────────────────────
_BOT_UA = [
    "bot","crawler","spider","scraper","wget","curl","python-requests",
    "httpx","aiohttp","scrapy","mechanize","selenium","playwright",
    "headlesschrome","phantomjs","slurp","baiduspider",
    "facebookexternalhit","twitterbot",
    "gptbot","chatgpt","claudebot","anthropic","ccbot","semrush",
    "ahrefsbot","mj12bot","dotbot","petalbot","bytespider",
    "zoominfobot","dataforseobot","blexbot","megaindex",
    "go-http-client","java/","libwww-perl","httpclient",
]
# 정상 검색엔진 봇 (홈페이지만 허용, 데이터 페이지는 차단)
_SEARCH_ENGINE_BOTS = ["googlebot", "bingbot", "yeti", "naverbot", "daumoa", "yandexbot"]

# IP당 분당 요청 기록 (최근 60초 타임스탬프 큐)
_rate_store: dict = collections.defaultdict(collections.deque)
_rate_store_cleanup_ts: float = 0.0  # 마지막 정리 시간
_honeypot_ips: set = set()  # 허니팟 걸린 IP 영구 차단
_RATE_LIMIT = 40  # 분당 최대 요청 수 (비인증 사용자)
_RATE_LIMIT_AUTH = 120  # 분당 최대 요청 수 (인증된 사용자)
_WHITELIST_PATHS = {"/static", "/data", "/robots.txt", "/favicon.ico", "/api/debug/", "/api/cron", "/api/img-proxy", "/api/self-collect", "/api/xpoz-collect"}
# 데이터 페이지 경로 (검색엔진 봇 차단 대상)
_DATA_PATHS = ["/influencers", "/api/", "/advertiser", "/export", "/collect",
               "/hashtags", "/settings", "/refresh", "/target-extract"]

def _get_client_ip(request: Request) -> str:
    """클라이언트 IP 추출 (프록시 지원)"""
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else ""

@app.middleware("http")
async def bot_protection(request: Request, call_next):
    global _rate_store_cleanup_ts
    path = request.url.path
    ip = _get_client_ip(request)
    ua = request.headers.get("user-agent", "").lower()

    # 허니팟에 걸린 IP 영구 차단
    if ip in _honeypot_ips:
        return Response("Forbidden", status_code=403)

    # 정적파일은 통과
    if any(path.startswith(p) for p in _WHITELIST_PATHS):
        return await call_next(request)

    # User-Agent가 비어있으면 차단 (정상 브라우저는 항상 UA를 보냄)
    if not ua or len(ua) < 10:
        log.warning(f"빈/짧은 UA 차단: {ip} UA='{ua}'")
        return Response("Forbidden", status_code=403)

    # 정상 검색엔진 봇: 홈페이지(/)만 허용, 데이터 페이지는 차단
    is_search_bot = any(sb in ua for sb in _SEARCH_ENGINE_BOTS)
    if is_search_bot:
        if path == "/" or path == "":
            return await call_next(request)
        else:
            log.info(f"검색엔진 봇 데이터 페이지 차단: {ip} UA={ua[:40]} path={path}")
            return Response("Forbidden", status_code=403)

    # User-Agent 봇 차단
    if any(b in ua for b in _BOT_UA):
        log.warning(f"봇 차단: {ip} UA={ua[:60]}")
        return Response("Forbidden", status_code=403)

    # 데이터센터/헤드리스 브라우저 힌트 차단
    if any(h in ua for h in _DATACENTER_UA_HINTS):
        log.warning(f"헤드리스 브라우저 차단: {ip} UA={ua[:60]}")
        return Response("Forbidden", status_code=403)

    # 클라이언트 JS 봇 탐지 쿠키 확인 (navigator.webdriver 등)
    if request.cookies.get("_bot") == "1":
        log.warning(f"JS 봇 탐지 쿠키 발견: {ip}")
        _honeypot_ips.add(ip)
        return Response("Forbidden", status_code=403)

    # TLS Fingerprinting (Vercel 환경에서만 작동)
    if not _check_tls_fingerprint(request):
        log.warning(f"TLS fingerprint 차단: {ip} JA4={request.headers.get('x-vercel-ja4-digest','')[:20]}")
        return Response("Forbidden", status_code=403)

    # Rate limiting
    now = time.time()
    has_session = bool(request.cookies.get("session_id") or request.cookies.get("adv_session_id"))
    limit = _RATE_LIMIT_AUTH if has_session else _RATE_LIMIT
    dq = _rate_store[ip]
    # 60초 이전 기록 제거
    while dq and dq[0] < now - 60:
        dq.popleft()
    if len(dq) >= limit:
        log.warning(f"Rate limit 초과: {ip} ({len(dq)}req/min, limit={limit})")
        return JSONResponse({"error": "Too many requests"}, status_code=429,
            headers={"Retry-After": "60", "X-RateLimit-Limit": str(limit)})
    dq.append(now)

    # 5분마다 오래된 IP 레코드 정리 (메모리 누수 방지)
    if now - _rate_store_cleanup_ts > 300:
        _rate_store_cleanup_ts = now
        stale = [k for k, v in _rate_store.items() if not v or v[-1] < now - 120]
        for k in stale:
            del _rate_store[k]

    response = await call_next(request)

    # 보안 헤더 추가
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # 데이터 페이지는 검색엔진 인덱싱 차단
    if any(path.startswith(dp) for dp in _DATA_PATHS):
        response.headers["X-Robots-Tag"] = "noindex, nofollow, noarchive"
        response.headers["Cache-Control"] = "private, no-store, no-cache, must-revalidate"

    return response


@app.get("/trap/data-feed")  # 허니팟: robots.txt에 Disallow되어 있지만 봇은 방문함
async def honeypot(request: Request):
    ip = _get_client_ip(request)
    _honeypot_ips.add(ip)
    log.warning(f"허니팟 감지! IP 영구 차단: {ip}")
    return Response("Not Found", status_code=404)

@app.get("/trap/admin-panel")  # 추가 허니팟
async def honeypot2(request: Request):
    ip = _get_client_ip(request)
    _honeypot_ips.add(ip)
    log.warning(f"허니팟2 감지! IP 영구 차단: {ip}")
    return Response("Not Found", status_code=404)


# ─── JS Challenge 토큰 (헤드리스 브라우저/스크래퍼 차단) ──────────
_JS_CHALLENGE_SECRET = os.getenv("JS_CHALLENGE_SECRET", secrets.token_hex(16))

def _generate_challenge_token(ip: str) -> str:
    """IP + 시간(10분 단위) 기반 HMAC 토큰 생성"""
    time_slot = str(int(time.time()) // 600)  # 10분마다 갱신
    msg = f"{ip}:{time_slot}".encode()
    return hmac.new(_JS_CHALLENGE_SECRET.encode(), msg, hashlib.sha256).hexdigest()[:32]

def _verify_challenge_token(ip: str, token: str) -> bool:
    """JS Challenge 토큰 검증 (현재 + 이전 10분 슬롯 허용)"""
    if not token:
        return False
    now_slot = int(time.time()) // 600
    for slot in [now_slot, now_slot - 1]:
        msg = f"{ip}:{slot}".encode()
        expected = hmac.new(_JS_CHALLENGE_SECRET.encode(), msg, hashlib.sha256).hexdigest()[:32]
        if hmac.compare_digest(token, expected):
            return True
    return False

@app.get("/api/challenge-token")
async def get_challenge_token(request: Request):
    """JS에서 호출하여 challenge 토큰을 받아가는 엔드포인트"""
    ip = _get_client_ip(request)
    return JSONResponse({"token": _generate_challenge_token(ip)})


# ─── 로그인 Brute-force 방지 ──────────────────────────────────────
_login_attempts: dict = collections.defaultdict(list)  # IP -> [timestamp, ...]
_LOGIN_MAX_ATTEMPTS = 5       # 15분 내 최대 시도
_LOGIN_LOCKOUT_SECS = 900     # 15분 잠금

def _check_login_allowed(ip: str) -> bool:
    """로그인 시도 허용 여부 확인"""
    now = time.time()
    attempts = _login_attempts[ip]
    # 15분 이전 기록 제거
    _login_attempts[ip] = [t for t in attempts if t > now - _LOGIN_LOCKOUT_SECS]
    return len(_login_attempts[ip]) < _LOGIN_MAX_ATTEMPTS

def _record_login_attempt(ip: str):
    """실패한 로그인 시도 기록"""
    _login_attempts[ip].append(time.time())


# ─── 데이터센터 IP 차단 (간이 ASN 체크) ──────────────────────────
_DATACENTER_UA_HINTS = [
    "headless", "phantomjs", "electron", "puppeteer",
    "splash", "nightmare", "casperjs", "slimerjs",
]


# ─── TLS Fingerprinting (Vercel JA4) ─────────────────────────────
# Vercel은 모든 요청에 x-vercel-ja4-digest 헤더를 자동 주입
# 알려진 봇/스크래퍼 TLS fingerprint를 차단
_KNOWN_BOT_JA4 = set()  # 탐지 시 동적으로 추가
_KNOWN_BROWSER_JA4_PREFIXES = [
    "t13d",  # TLS 1.3 (Chrome, Firefox, Safari, Edge)
    "t12d",  # TLS 1.2 (구형 브라우저)
]

def _check_tls_fingerprint(request: Request) -> bool:
    """Vercel JA4 fingerprint로 비브라우저 클라이언트 탐지.
    Vercel 환경이 아니면 (로컬 개발) 통과."""
    ja4 = request.headers.get("x-vercel-ja4-digest", "")
    if not ja4:
        return True  # 로컬 개발 환경 또는 JA4 미지원 → 통과
    if ja4 in _KNOWN_BOT_JA4:
        return False
    # TLS 1.3 브라우저 fingerprint 패턴 확인
    # JA4 형식: t13d1516h2_8daaf6152771_... (TLS버전_ciphers_extensions)
    if any(ja4.startswith(p) for p in _KNOWN_BROWSER_JA4_PREFIXES):
        return True
    # 알 수 없는 JA4 → 로그 기록 후 통과 (false positive 방지, 모니터링용)
    log.info(f"미확인 JA4 fingerprint: {ja4[:30]}")
    return True


# ─── 워터마킹 (데이터 유출 추적) ─────────────────────────────────
# Zero-Width 유니코드 문자로 사용자 ID를 텍스트에 인코딩
_ZW_CHARS = ['\u200b', '\u200c', '\u200d', '\ufeff']  # ZWS, ZWNJ, ZWJ, BOM

def _encode_watermark(user_id: str) -> str:
    """사용자 ID를 보이지 않는 Zero-Width 문자열로 인코딩"""
    # user_id의 각 문자를 2비트씩 4개의 ZW 문자로 매핑
    bits = ''.join(format(ord(c), '08b') for c in user_id[:8])  # 최대 8자
    result = []
    for i in range(0, len(bits), 2):
        idx = int(bits[i:i+2], 2)  # 0~3
        result.append(_ZW_CHARS[idx])
    return ''.join(result)

def _watermark_text(text: str, user_id: str) -> str:
    """텍스트에 보이지 않는 워터마크 삽입"""
    if not text or not user_id:
        return text
    wm = _encode_watermark(user_id)
    # 텍스트 중간에 워터마크 삽입 (첫 번째 공백 뒤)
    idx = text.find(' ')
    if idx > 0:
        return text[:idx] + wm + text[idx:]
    return text + wm

def _watermark_number(value: int, user_id: str) -> int:
    """숫자에 미세한 워터마크 (±0.1% 이내 변동, 사용자별 고유)"""
    if not value or value < 100:
        return value
    # user_id 기반 결정론적 오프셋 (-0.1% ~ +0.1%)
    h = int(hashlib.md5(f"{user_id}:{value}".encode()).hexdigest()[:4], 16)
    offset_pct = (h % 200 - 100) / 100000  # -0.001 ~ +0.001
    return int(value * (1 + offset_pct))


# ─── Proof-of-Work (PoW) ─────────────────────────────────────────
# 클라이언트가 SHA-256 해시를 찾아야 데이터 접근 가능
_POW_DIFFICULTY = 4  # 해시 앞 4자리가 '0000'이어야 함 (평균 ~65,000번 시도, ~200ms)
_POW_SECRET = os.getenv("POW_SECRET", secrets.token_hex(8))
_pow_used_nonces: dict = {}  # nonce -> expiry timestamp (재사용 방지)
_pow_cleanup_ts: float = 0.0

def _generate_pow_challenge() -> dict:
    """PoW 챌린지 생성"""
    challenge_id = secrets.token_hex(8)
    timestamp = int(time.time())
    return {
        "challenge": challenge_id,
        "timestamp": timestamp,
        "difficulty": _POW_DIFFICULTY,
    }

def _verify_pow(challenge: str, timestamp: int, nonce: str) -> bool:
    """PoW 솔루션 검증"""
    global _pow_cleanup_ts
    now = time.time()
    # 타임스탬프 유효성 (5분 이내)
    if abs(now - timestamp) > 300:
        return False
    # nonce 재사용 방지
    nonce_key = f"{challenge}:{nonce}"
    if nonce_key in _pow_used_nonces:
        return False
    # 해시 검증
    data = f"{challenge}:{timestamp}:{nonce}:{_POW_SECRET}".encode()
    h = hashlib.sha256(data).hexdigest()
    if not h.startswith('0' * _POW_DIFFICULTY):
        return False
    # nonce 등록 (5분 TTL)
    _pow_used_nonces[nonce_key] = now + 300
    # 10분마다 만료된 nonce 정리
    if now - _pow_cleanup_ts > 600:
        _pow_cleanup_ts = now
        expired = [k for k, v in _pow_used_nonces.items() if v < now]
        for k in expired:
            del _pow_used_nonces[k]
    return True

@app.get("/api/pow-challenge")
async def pow_challenge(request: Request):
    """PoW 챌린지 발급"""
    return JSONResponse(_generate_pow_challenge())

@app.post("/api/pow-verify")
async def pow_verify(request: Request):
    """PoW 솔루션 검증 → 성공 시 데이터 접근 토큰 발급"""
    try:
        body = await request.json()
        challenge = body.get("challenge", "")
        timestamp = body.get("timestamp", 0)
        nonce = body.get("nonce", "")
        if _verify_pow(challenge, timestamp, nonce):
            ip = _get_client_ip(request)
            token = _generate_challenge_token(ip)
            return JSONResponse({"ok": True, "token": token})
        return JSONResponse({"ok": False, "error": "Invalid solution"}, status_code=400)
    except Exception:
        return JSONResponse({"ok": False, "error": "Bad request"}, status_code=400)


@app.get("/robots.txt")
async def robots():
    content = """# 정상 검색엔진: 홈페이지만 허용
User-agent: Googlebot
Allow: /$
Disallow: /

User-agent: Yeti
Allow: /$
Disallow: /

User-agent: Bingbot
Allow: /$
Disallow: /

User-agent: Naverbot
Allow: /$
Disallow: /

User-agent: DaumOa
Allow: /$
Disallow: /

User-agent: Yandexbot
Allow: /$
Disallow: /

# AI/스크래핑 봇 전면 차단
User-agent: GPTBot
Disallow: /
User-agent: ChatGPT-User
Disallow: /
User-agent: CCBot
Disallow: /
User-agent: anthropic-ai
Disallow: /
User-agent: ClaudeBot
Disallow: /
User-agent: Bytespider
Disallow: /
User-agent: SemrushBot
Disallow: /
User-agent: AhrefsBot
Disallow: /
User-agent: MJ12bot
Disallow: /
User-agent: DataForSeoBot
Disallow: /
User-agent: PetalBot
Disallow: /

# 기타 모든 봇 전면 차단
User-agent: *
Disallow: /
Disallow: /api/
Disallow: /influencers/
Disallow: /advertiser/
Disallow: /export/
Disallow: /collect/
Disallow: /trap/
Disallow: /settings/
"""
    return Response(content, media_type="text/plain")

# Supabase 모드(프로덕션)이면 /tmp 사용 (Vercel 읽기전용 FS)
_IS_PROD = bool(os.environ.get("SUPABASE_KEY"))
_LOCAL_DATA = os.path.join(os.path.dirname(__file__), "data")
DATA_DIR = "/tmp/data" if _IS_PROD else _LOCAL_DATA
try:
    os.makedirs(os.path.join(DATA_DIR, "profile_pics"), exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, "posts"), exist_ok=True)
except OSError:
    DATA_DIR = "/tmp/data"
    os.makedirs(os.path.join(DATA_DIR, "profile_pics"), exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, "posts"), exist_ok=True)

if not _IS_PROD and os.path.isdir(DATA_DIR):
    app.mount("/data", StaticFiles(directory=DATA_DIR), name="data")
_STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
os.makedirs(_STATIC_DIR, exist_ok=True)
app.mount("/static", StaticFiles(directory=_STATIC_DIR), name="static")

templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "templates"))
def _dt_filter(t):
    if not t: return "-"
    try:
        return datetime.fromtimestamp(float(t), tz=timezone(timedelta(hours=9))).strftime("%m/%d %H:%M")
    except (ValueError, TypeError, OSError):
        # ISO 문자열 처리
        try:
            s = str(t)
            if "T" in s:
                from datetime import datetime as _dtc
                dt = _dtc.fromisoformat(s.replace("Z", "+00:00"))
                return dt.astimezone(timezone(timedelta(hours=9))).strftime("%m/%d %H:%M")
        except Exception:
            pass
    return str(t)[:16] if t else "-"
templates.env.filters["dt"] = _dt_filter
def _fmtn(n):
    v = int(n or 0)
    if v <= 0: return "0"
    if v >= 1_000_000: return f"{v/1_000_000:.1f}".rstrip('0').rstrip('.') + "M"
    if v >= 1_000: return f"{v/1_000:.1f}".rstrip('0').rstrip('.') + "K"
    return str(v)
templates.env.filters["fmtn"] = _fmtn
def _comma(n):
    try:
        if n is None or n == '' or n == '?': return "0"
        return f"{int(float(str(n).replace(',','')))  :,}"
    except: return str(n) if n else "0"
templates.env.filters["comma"] = _comma
def _fromjson(s):
    try: return json.loads(s) if isinstance(s, str) else s
    except: return {}
templates.env.filters["fromjson"] = _fromjson

def _pic(inf, size=128):
    """프로필 사진 URL: Supabase Storage > CDN > fallback"""
    _sb_url = os.environ.get("SUPABASE_URL", "https://ysqnixgdpltguatvjjcb.supabase.co")
    _SB_STORAGE = f"{_sb_url}/storage/v1/object/public/profile-pics/"
    if isinstance(inf, dict):
        local = inf.get("profile_pic_local") or ""
        cdn = inf.get("profile_pic_url") or ""
        pk = inf.get("pk") or ""
        name = inf.get("username") or inf.get("full_name") or "U"
    else:
        local = cdn = pk = ""
        name = str(inf) if inf else "U"
    # 1) profile_pic_local이 Supabase 전체 URL이면 그대로
    if local and local.startswith("http"):
        return local
    # 2) pk가 있으면 Supabase Storage에서 {pk}.jpg 시도 (CDN보다 안정적)
    if pk:
        return f"{_SB_STORAGE}{pk}.jpg"
    # 3) CDN URL 폴백
    if cdn:
        return cdn
    from urllib.parse import quote
    return f"https://ui-avatars.com/api/?name={quote(name[:2])}&background=6366f1&color=fff&size={size}"
templates.env.filters["pic"] = _pic

@app.get("/api/img-proxy")
def img_proxy(url: str = ""):
    """Instagram CDN 이미지 프록시 — referrer 차단 및 만료 URL 우회"""
    if not url or not url.startswith("https://"):
        return Response(b"", status_code=400)
    import requests as _req
    try:
        r = _req.get(url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200 and len(r.content) > 500:
            ct = r.headers.get("content-type", "image/jpeg")
            return Response(r.content, media_type=ct,
                            headers={"Cache-Control": "public, max-age=86400"})
    except Exception:
        pass
    return Response(b"", status_code=404)


def _safe_cd(fname: str) -> str:
    """Content-Disposition 헤더용 RFC 5987 인코딩"""
    from urllib.parse import quote
    ascii_name = fname.encode("ascii", "ignore").decode("ascii") or "download"
    encoded = quote(fname, safe="")
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded}"

# ── 인증 (관리자) - JWT 기반 (Vercel 서버리스 호환) ──────────
from jose import jwt as _jwt
JWT_SECRET = os.getenv("SECRET_KEY", "supertag-secret-key-2026-supers")
JWT_ALG = "HS256"
JWT_EXP_HOURS = 24 * 7  # 7일

ADMIN_PW_HASH = bcrypt.hashpw(
    os.getenv("ADMIN_PASSWORD", "admin").encode(), bcrypt.gensalt()
)
ADMIN_USER = os.getenv("ADMIN_USERNAME", "admin")
EXTENSION_API_KEY = os.getenv("EXTENSION_API_KEY", "supertag-ext-key")

def _check_ext_key(request: Request) -> bool:
    return request.headers.get("X-Api-Key") == EXTENSION_API_KEY

def _make_jwt(payload: dict, hours: int = JWT_EXP_HOURS) -> str:
    import datetime
    data = dict(payload)
    data["exp"] = datetime.datetime.utcnow() + datetime.timedelta(hours=hours)
    return _jwt.encode(data, JWT_SECRET, algorithm=JWT_ALG)

def _decode_jwt(token: str) -> dict | None:
    try:
        return _jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except Exception:
        return None

# Fallback in-memory sessions for local dev (not used in prod)
sessions: dict = {}
adv_sessions: dict = {}

INSTA_CFG = {
    "username": os.getenv("INSTA_USERNAME", "jannat160304"),
    "password": os.getenv("INSTA_PASSWORD", "jug@575"),
    "totp":     os.getenv("INSTA_TOTP", "YQ754N2HTC7IDAT5BQIPNA5RHQA75JFY"),
}

# 구독 플랜 정의
PLANS = {
    "free":       {"name": "무료",       "per_hashtag": 100,   "daily_limit": 500,  "max_hashtags": 5,  "price_1m": 0,    "price_3m": 0},
    "starter":    {"name": "스타터",     "per_hashtag": 1000,  "daily_limit": 5000, "max_hashtags": 10, "price_1m": 10000, "price_3m": 27000},
    "pro":        {"name": "프로",       "per_hashtag": 5000,  "daily_limit": 20000,"max_hashtags": 30, "price_1m": 30000, "price_3m": 81000},
    "enterprise": {"name": "엔터프라이즈","per_hashtag": 10000, "daily_limit": 50000,"max_hashtags": 100,"price_1m": 50000, "price_3m": 135000},
}


def get_user(session_id: Optional[str] = None):
    """session_id는 실제로 JWT token (cookie name 유지)"""
    if not session_id:
        return None
    payload = _decode_jwt(session_id)
    if payload and payload.get("role") == "admin":
        return {"username": payload.get("username", "admin")}
    # Fallback: in-memory (local dev)
    return sessions.get(session_id)

def get_adv_user(session_id: Optional[str] = None):
    """adv_session_id는 JWT token"""
    if not session_id:
        return None
    payload = _decode_jwt(session_id)
    if payload and payload.get("role") == "advertiser":
        adv_id = payload.get("adv_id")
        if adv_id:
            # Supabase에서 최신 정보 가져오기
            from database import _USE_SUPABASE
            if _USE_SUPABASE:
                rows = __import__("database")._sb_get(
                    "insta_advertiser_accounts",
                    {"id": f"eq.{adv_id}", "limit": "1"}
                )
                return rows[0] if rows else None
        return payload
    return adv_sessions.get(session_id)

def require_admin(session_id: Optional[str]):
    user = get_user(session_id)
    if not user:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    return user


@app.on_event("startup")
def startup():
    init_db()


# ═══════════════════════════════════════════════════════
# 공개 메인 페이지 (로그인 없이)
# ═══════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
def public_home(
    request: Request,
    session_id: Optional[str] = Cookie(default=None)
):
    stats = get_public_stats()
    user = get_user(session_id)
    return templates.TemplateResponse("home.html", {
        "request": request, "stats": stats, "user": user,
    })


@app.get("/api/public")
def api_public(
    request: Request,
    page: int = Query(1, gt=0),
    sort: str = "follower_count",
    q: str = "",
    min_f: int = 0,
    max_f: int = 0,
    category: str = "",
    hashtag: str = "",
    public_only: bool = False,
    no_biz: bool = False,       # 비즈니스 계정 제외
    biz_only: bool = False,     # 비즈니스 계정만
    verified_only: bool = False, # 인증 계정만
    session_id: Optional[str] = Cookie(default=None),
    adv_session_id: Optional[str] = Cookie(default=None),
):
    """공개 인플루언서 목록 JSON API - JS 렌더링용
    비인증 사용자: 제한된 필드만 반환, 페이지 제한, JS Challenge 필수
    인증 사용자: 전체 필드 반환
    """
    is_authorized = bool(get_user(session_id) or get_adv_user(adv_session_id))
    # 비인증 사용자: JS Challenge 토큰 검증 (헤드리스 크롤러 차단)
    if not is_authorized:
        challenge = request.headers.get("X-Challenge-Token", "")
        ip = _get_client_ip(request)
        if not _verify_challenge_token(ip, challenge):
            return JSONResponse({"error": "Challenge token required", "total": 0, "rows": []}, status_code=403)
    per_page = 30 if is_authorized else 10
    # 비인증 사용자는 최대 3페이지까지만 허용 (데이터 대량 수집 방지)
    if not is_authorized and page > 3:
        return JSONResponse({"error": "로그인이 필요합니다", "total": 0, "rows": []}, status_code=401)
    total, rows = get_public_influencers(page=page, per_page=per_page, sort=sort,
                                         q=q, min_f=min_f, max_f=max_f,
                                         category=category, hashtag=hashtag,
                                         public_only=public_only, no_biz=no_biz,
                                         biz_only=biz_only, verified_only=verified_only)
    # 최종 안전망: 숨김/밴 계정 재필터링
    from database import get_hidden_pks, get_banned_pks
    _excluded = {str(p) for p in (get_hidden_pks() | get_banned_pks())}
    if _excluded:
        before = len(rows)
        rows = [r for r in rows if str(r.get("pk","")) not in _excluded]
        total = max(0, total - (before - len(rows)))
    total_pages = max(1, (total + per_page - 1) // per_page)
    def _g(r, k, d=0):
        return r[k] if isinstance(r, dict) else getattr(r, k, d)
    # 워터마킹용 사용자 ID 추출
    _wm_id = ""
    if is_authorized:
        u = get_user(session_id)
        a = get_adv_user(adv_session_id)
        _wm_id = (u.get("username","") if u else "") or (a.get("username","") if a else "") or _get_client_ip(request)

    result = []
    for r in rows:
        if is_authorized:
            # 인증된 사용자: 전체 데이터 (워터마크 포함)
            fc = int(_g(r,"follower_count",0))
            result.append({
                "pk": _g(r,"pk",""),
                "username": _g(r,"username",""),
                "full_name": _watermark_text(str(_g(r,"full_name","")), _wm_id),
                "biography": _watermark_text(str(_g(r,"biography","")), _wm_id),
                "follower_count": _watermark_number(fc, _wm_id),
                "engagement_rate": _g(r,"engagement_rate",0),
                "avg_likes": _g(r,"avg_likes",0),
                "avg_comments": _g(r,"avg_comments",0),
                "avg_reel_views": _g(r,"avg_reel_views",0),
                "avg_reel_likes": _g(r,"avg_reel_likes",0),
                "avg_reel_comments": _g(r,"avg_reel_comments",0),
                "avg_feed_likes": _g(r,"avg_feed_likes",0),
                "avg_feed_comments": _g(r,"avg_feed_comments",0),
                "is_verified": _g(r,"is_verified",False),
                "is_business": _g(r,"is_business",False),
                "is_private": bool(_g(r,"is_private",False)),
                "category": _g(r,"category",""),
                "hashtags": _g(r,"hashtags",""),
                "profile_pic_url": _pic(r),
                "profile_pic_local": _pic(r),
            })
        else:
            # 비인증 사용자: 제한된 필드만 (상세 통계 숨김)
            result.append({
                "pk": _g(r,"pk",""),
                "username": _g(r,"username",""),
                "full_name": _g(r,"full_name",""),
                "follower_count": _g(r,"follower_count",0),
                "is_verified": _g(r,"is_verified",False),
                "is_business": _g(r,"is_business",False),
                "category": _g(r,"category",""),
                "profile_pic_url": _pic(r),
                "profile_pic_local": _pic(r),
            })
    # 마지막 업데이트 날짜
    last_updated = ""
    try:
        from database import get_collect_jobs
        _recent = get_collect_jobs(limit=1)
        if _recent:
            _ft = _recent[0].get("finished_at") or _recent[0].get("started_at")
            if _ft and isinstance(_ft, str) and "T" in _ft:
                from datetime import datetime as _dt2
                _d = _dt2.fromisoformat(_ft)
                last_updated = _d.strftime("%Y.%m.%d %H:%M")
            elif _ft:
                from datetime import datetime as _dt2, timezone as _tz2, timedelta as _td2
                _d = _dt2.fromtimestamp(float(_ft), tz=_tz2(_td2(hours=9)))
                last_updated = _d.strftime("%Y.%m.%d %H:%M")
    except Exception:
        pass
    return {"total": total, "total_pages": total_pages, "page": page, "rows": result, "last_updated": last_updated}


# ═══════════════════════════════════════════════════════
# 관리자 로그인
# ═══════════════════════════════════════════════════════

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    ip = _get_client_ip(request)
    if not _check_login_allowed(ip):
        log.warning(f"로그인 brute-force 차단: {ip}")
        return templates.TemplateResponse("login.html", {
            "request": request, "error": "너무 많은 로그인 시도입니다. 15분 후 다시 시도해주세요."
        }, status_code=429)
    if username == ADMIN_USER and bcrypt.checkpw(password.encode(), ADMIN_PW_HASH):
        _login_attempts.pop(ip, None)  # 성공 시 기록 초기화
        token = _make_jwt({"role": "admin", "username": username})
        res = RedirectResponse("/influencers", status_code=302)
        res.set_cookie("session_id", token, httponly=True, max_age=604800,
                        samesite="lax", secure=_IS_PROD)
        return res
    _record_login_attempt(ip)
    remaining = _LOGIN_MAX_ATTEMPTS - len(_login_attempts.get(ip, []))
    return templates.TemplateResponse("login.html", {
        "request": request, "error": f"아이디 또는 비밀번호가 올바르지 않습니다. (남은 시도: {remaining}회)"
    }, status_code=400)

@app.get("/logout")
def logout(session_id: Optional[str] = Cookie(default=None)):
    if session_id in sessions: del sessions[session_id]
    res = RedirectResponse("/login", 302)
    res.delete_cookie("session_id")
    return res


# ═══════════════════════════════════════════════════════
# 인플루언서 목록 (관리자 전용)
# ═══════════════════════════════════════════════════════

@app.get("/influencers", response_class=HTMLResponse)
def influencers(
    request: Request,
    q: str = "",
    hashtag: str = "",
    min_f: Optional[str] = None,
    max_f: Optional[str] = None,
    verified: int = 0,
    public_only: int = 0,
    main_category: str = "",
    can_live: int = 0,
    only_approved: int = 0,
    has_pet: int = 0,
    is_married: int = 0,
    has_kids: int = 0,
    has_car: int = 0,
    is_visual: int = 0,
    has_url: int = 0,
    url_domain: str = "",
    sort: str = "follower_count",
    order: str = "desc",
    page: int = Query(1, gt=0),
    per_page: int = Query(50, gt=0),
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id)
    if not user:
        return RedirectResponse(f"/login?next=/influencers", 302)

    _min_f = int(min_f) if min_f and min_f.isdigit() else None
    _max_f = int(max_f) if max_f and max_f.isdigit() else None

    total, rows = get_influencers(
        keyword=q, hashtag_filter=hashtag,
        min_f=_min_f, max_f=_max_f,
        only_verified=bool(verified), exclude_private=bool(public_only),
        main_category=main_category,
        can_live=bool(can_live), only_approved=bool(only_approved),
        has_pet=bool(has_pet), is_married=bool(is_married),
        has_kids=bool(has_kids), has_car=bool(has_car),
        is_visual=bool(is_visual),
        has_url=bool(has_url), url_domain=url_domain,
        sort=sort, order=order, page=page, per_page=per_page
    )
    stats = get_stats()
    refresh = get_refresh_status()
    total_pages = max(1, (total + per_page - 1) // per_page)
    # 검색어가 있고 결과가 없으면 즉시 수집 제안
    instant_collect_target = ""
    if q and total == 0 and re.match(r'^[a-zA-Z0-9._]+$', q):
        instant_collect_target = q
    return templates.TemplateResponse("influencers.html", {
        "request": request, "user": user,
        "rows": rows, "total": total, "total_pages": total_pages,
        "page": page, "per_page": per_page,
        "q": q, "hashtag": hashtag,
        "min_f": min_f, "max_f": max_f,
        "verified": verified, "public_only": public_only,
        "main_category": main_category,
        "can_live": can_live, "only_approved": only_approved,
        "has_pet": has_pet, "is_married": is_married, "has_kids": has_kids, "has_car": has_car,
        "is_visual": is_visual, "has_url": has_url, "url_domain": url_domain,
        "sort": sort, "order": order,
        "stats": stats, "refresh": refresh,
        "instant_collect_target": instant_collect_target,
        "url_stats": get_url_stats(),
    })


@app.get("/api/influencers")
def api_influencers(
    q: str = "", hashtag: str = "",
    min_f: Optional[str] = None, max_f: Optional[str] = None,
    verified: int = 0, public_only: int = 0,
    main_category: str = "", can_live: int = 0, only_approved: int = 0,
    has_pet: int = 0, is_married: int = 0, has_kids: int = 0, has_car: int = 0,
    is_visual: int = 0,
    sort: str = "follower_count", order: str = "desc",
    page: int = Query(1, gt=0), per_page: int = Query(50, gt=0),
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    _min_f = int(min_f) if min_f and min_f.isdigit() else None
    _max_f = int(max_f) if max_f and max_f.isdigit() else None
    total, rows = get_influencers(
        keyword=q, hashtag_filter=hashtag, min_f=_min_f, max_f=_max_f,
        only_verified=bool(verified), exclude_private=bool(public_only),
        main_category=main_category, can_live=bool(can_live), only_approved=bool(only_approved),
        has_pet=bool(has_pet), is_married=bool(is_married),
        has_kids=bool(has_kids), has_car=bool(has_car), is_visual=bool(is_visual),
        sort=sort, order=order, page=page, per_page=per_page
    )
    # 최종 안전망: 숨김/밴 계정 재필터링
    from database import get_hidden_pks, get_banned_pks
    _excluded = {str(p) for p in (get_hidden_pks() | get_banned_pks())}
    if _excluded:
        rows = [r for r in rows if str(r.get("pk","")) not in _excluded]
    for r in rows:
        r["pic_url"] = _pic(r)
    return JSONResponse({"total": total, "rows": rows, "page": page, "per_page": per_page})


# ═══════════════════════════════════════════════════════
# 즉시 수집 (계정명으로 실시간 크롤링)
# ═══════════════════════════════════════════════════════

@app.post("/influencers/instant-collect")
async def instant_collect(
    background_tasks: BackgroundTasks,
    username: str = Form(...),
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "로그인 필요"}, 401)
    username = username.strip().lstrip("@")
    if not re.match(r'^[a-zA-Z0-9._]+$', username):
        return JSONResponse({"error": "올바르지 않은 계정명"}, 400)

    from crawler import crawl_single_user
    # 백그라운드로 실행
    background_tasks.add_task(_run_instant_collect, username)
    return JSONResponse({"ok": True, "message": f"@{username} 수집 시작"})


def _run_instant_collect(username: str):
    from crawler import crawl_single_user
    crawl_single_user(username)


# ═══════════════════════════════════════════════════════
# 인플루언서 상세
# ═══════════════════════════════════════════════════════

@app.get("/influencers/{pk}", response_class=HTMLResponse)
def influencer_detail(
    pk: str, request: Request,
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id)
    if not user:
        return RedirectResponse(f"/login", 302)
    inf = get_influencer(pk)
    if not inf:
        raise HTTPException(404, "인플루언서를 찾을 수 없습니다")
    manual = get_manual(pk)
    posts = get_influencer_posts(pk)

    # JSON 파싱
    for key in ["top_posts_likes", "top_posts_comments", "top_reels_views"]:
        try:
            inf[key] = json.loads(inf.get(key) or "[]")
        except:
            inf[key] = []

    # top_hashtags 파싱 (구형: ["뷰티",...] 또는 신형: [{"tag":"뷰티","count":5},...])
    try:
        raw = json.loads(inf.get("top_hashtags") or "[]")
        if raw and isinstance(raw[0], str):
            inf["top_hashtags"] = [{"tag": t, "count": 1} for t in raw]
        else:
            inf["top_hashtags"] = raw
    except:
        inf["top_hashtags"] = []

    # ── 해시태그 활동 유형 분석 ──
    activity_analysis = _analyze_activity(inf.get("top_hashtags", []))

    # posts DB에 있으면 top posts도 posts 기반으로 재계산
    if posts:
        reels_p = [p for p in posts if p.get("post_type") == "reel"]
        feeds_p = [p for p in posts if p.get("post_type") != "reel"]
        all_p = list(posts)
        reels_p.sort(key=lambda x: x.get("views", 0) or 0, reverse=True)
        feeds_p.sort(key=lambda x: x.get("likes", 0) or 0, reverse=True)
        if reels_p:
            inf["top_reels_views"] = [
                {"url": p["post_url"], "views": p.get("views", 0),
                 "thumbnail": p.get("thumbnail_url", "") or p.get("thumbnail_local", "")}
                for p in reels_p[:5]
            ]
        if feeds_p:
            inf["top_posts_likes"] = [
                {"url": p["post_url"], "likes": p.get("likes", 0),
                 "thumbnail": p.get("thumbnail_url", "") or p.get("thumbnail_local", "")}
                for p in feeds_p[:5]
            ]
        # 댓글순 TOP 5
        comments_sorted = sorted(all_p, key=lambda x: x.get("comments", 0) or 0, reverse=True)
        inf["top_posts_comments"] = [
            {"url": p["post_url"], "comments": p.get("comments", 0), "likes": p.get("likes", 0),
             "post_type": p.get("post_type", ""), "views": p.get("views", 0),
             "thumbnail": p.get("thumbnail_url", "") or p.get("thumbnail_local", "")}
            for p in comments_sorted[:5]
        ]

    recent_reels = get_influencer_reels(pk, sort="recent", limit=12)
    popular_reels = get_influencer_reels(pk, sort="popular", limit=12)

    return templates.TemplateResponse("influencer_detail.html", {
        "request": request, "user": user,
        "inf": inf, "manual": manual, "posts": posts,
        "activity": activity_analysis,
        "recent_reels": recent_reels,
        "popular_reels": popular_reels,
    })


# ═══════════════════════════════════════════════════════
# 수동 입력 편집
# ═══════════════════════════════════════════════════════

@app.get("/influencers/{pk}/edit", response_class=HTMLResponse)
def edit_page(pk: str, request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    inf = get_influencer(pk)
    if not inf: raise HTTPException(404)
    manual = get_manual(pk)
    return templates.TemplateResponse("influencer_edit.html", {
        "request": request, "user": user, "inf": inf, "manual": manual,
    })

@app.post("/influencers/{pk}/edit")
def edit_save(
    pk: str,
    session_id: Optional[str] = Cookie(default=None),
    contact_name:  str = Form(default=""),
    contact_kakao: str = Form(default=""),
    contact_line:  str = Form(default=""),
    contact_email: str = Form(default=""),
    contact_phone: str = Form(default=""),
    can_live:      int = Form(default=0),
    live_platforms: str = Form(default=""),
    live_price:    int = Form(default=0),
    feed_price:    int = Form(default=0),
    reel_price:    int = Form(default=0),
    story_price:   int = Form(default=0),
    bundle_price:  int = Form(default=0),
    main_category: str = Form(default=""),
    sub_categories: str = Form(default=""),
    target_gender: str = Form(default=""),
    target_age:    str = Form(default=""),
    target_region: str = Form(default=""),
    collab_types:  str = Form(default=""),
    past_brands:   str = Form(default=""),
    quality_score: int = Form(default=0),
    notes:         str = Form(default=""),
    is_approved:   int = Form(default=0),
    has_pet:       int = Form(default=0),
    is_married:    int = Form(default=0),
    has_kids:      int = Form(default=0),
    has_car:       int = Form(default=0),
    pet_type:      str = Form(default=""),
    kids_age:      str = Form(default=""),
    is_brand:      int = Form(default=0),
    is_visual:     int = Form(default=0),
    face_exposed:  int = Form(default=0),
    tiktok_url:    str = Form(default=""),
    youtube_url:   str = Form(default=""),
    facebook_url:  str = Form(default=""),
    threads_url:   str = Form(default=""),
    tiktok_followers:   int = Form(default=0),
    youtube_subscribers: int = Form(default=0),
    agency:        str = Form(default=""),
):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    save_manual(pk, {
        "contact_name": contact_name, "contact_kakao": contact_kakao,
        "contact_line": contact_line, "contact_email": contact_email,
        "contact_phone": contact_phone,
        "can_live": can_live, "live_platforms": live_platforms, "live_price": live_price,
        "feed_price": feed_price, "reel_price": reel_price,
        "story_price": story_price, "bundle_price": bundle_price,
        "main_category": main_category, "sub_categories": sub_categories,
        "target_gender": target_gender, "target_age": target_age, "target_region": target_region,
        "collab_types": collab_types, "past_brands": past_brands,
        "quality_score": quality_score, "notes": notes, "is_approved": is_approved,
        "approved_at": time.time() if is_approved else 0,
        "has_pet": has_pet, "is_married": is_married, "has_kids": has_kids, "has_car": has_car,
        "pet_type": pet_type, "kids_age": kids_age,
        "is_brand": is_brand, "is_visual": is_visual, "face_exposed": face_exposed,
        "tiktok_url": tiktok_url, "youtube_url": youtube_url,
        "facebook_url": facebook_url, "threads_url": threads_url,
        "tiktok_followers": tiktok_followers, "youtube_subscribers": youtube_subscribers,
        "agency": agency,
    })
    return RedirectResponse(f"/influencers/{pk}?saved=1", 302)


# ═══════════════════════════════════════════════════════
# 단일 인플루언서 상세 수집 (관리자)
# ═══════════════════════════════════════════════════════

@app.post("/influencers/{pk}/mark-brand")
def mark_brand(pk: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import save_manual
    save_manual(pk, {"is_brand": 1})
    return JSONResponse({"ok": True})


@app.post("/influencers/ban-bulk")
def ban_bulk(session_id: Optional[str] = Cookie(default=None),
             pks: str = Form(default=""), reason: str = Form(default="수동 밴")):
    """여러 인플루언서를 한 번에 밴"""
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import ban_influencer
    pk_list = [p.strip() for p in pks.split(",") if p.strip()]
    done = 0
    for pk in pk_list:
        try:
            ban_influencer(pk, reason)
            done += 1
        except Exception:
            pass
    return JSONResponse({"ok": True, "count": done, "total": len(pk_list)})


@app.post("/influencers/{pk}/ban")
def ban_one(pk: str, reason: str = Form(default="스팸/광고 계정"),
            session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import ban_influencer
    ban_influencer(pk, reason)
    return JSONResponse({"ok": True})


@app.post("/influencers/{pk}/delete")
def delete_influencer_route(pk: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import delete_influencer
    delete_influencer(pk)
    return JSONResponse({"ok": True})


@app.post("/influencers/{pk}/hide")
def hide_one(pk: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import hide_influencer
    hide_influencer(pk)
    return JSONResponse({"ok": True})


@app.post("/influencers/{pk}/unhide")
def unhide_one(pk: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import unhide_influencer
    unhide_influencer(pk)
    return JSONResponse({"ok": True})


@app.post("/influencers/{pk}/unban")
def unban_one(pk: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import unban_influencer
    unban_influencer(pk)
    return JSONResponse({"ok": True})

@app.post("/influencers/{pk}/ban-reason")
def update_ban_reason(pk: str, reason: str = Form(default=""),
                      session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import save_manual
    save_manual(pk, {"ban_reason": reason})
    return JSONResponse({"ok": True})


@app.get("/banned", response_class=HTMLResponse)
def banned_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    from database import get_banned_list
    banned = get_banned_list()
    return templates.TemplateResponse("banned.html", {
        "request": request, "user": user, "banned": banned,
    })

@app.get("/hidden", response_class=HTMLResponse)
def hidden_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    from database import get_hidden_list
    hidden = get_hidden_list()
    return templates.TemplateResponse("hidden.html", {
        "request": request, "user": user, "hidden": hidden,
    })

_last_errors = []   # 최근 에러 저장 (디버그용)

@app.get("/api/debug/version")
def debug_version():
    commit = os.environ.get("RAILWAY_GIT_COMMIT_SHA", "")[:7] or "unknown"
    return JSONResponse({"commit": commit, "deployed": True, "errors": _last_errors[-5:]})


@app.get("/api/debug/media-raw")
def debug_media_raw(username: str = ""):
    """HikerAPI 원본 미디어 응답 구조 확인 (첫 2개 게시물의 키/뷰 관련 필드)."""
    from crawler import _hiker_user_info, _hiker_user_medias
    if not username:
        return JSONResponse({"error": "?username= 필요"})
    info = _hiker_user_info(username)
    if not info:
        return JSONResponse({"error": "유저 조회 실패"})
    medias = _hiker_user_medias(str(info["pk"]), amount=3)
    if not medias:
        return JSONResponse({"error": "게시물 조회 실패"})
    result = []
    view_keys = ["view_count","play_count","video_play_count","video_view_count",
                 "ig_play_count","fb_play_count"]
    for m in medias[:3]:
        entry = {
            "all_keys": sorted(m.keys()) if isinstance(m, dict) else [],
            "media_type": m.get("media_type"),
            "product_type": m.get("product_type", ""),
            "code": m.get("code", ""),
        }
        for vk in view_keys:
            entry[f"field_{vk}"] = m.get(vk)
        # clips_metadata
        cm = m.get("clips_metadata")
        if isinstance(cm, dict):
            entry["clips_metadata_keys"] = sorted(cm.keys())
            entry["clips_metadata_play_count"] = cm.get("play_count")
        # thumbnail 관련
        entry["has_thumbnail_url"] = bool(m.get("thumbnail_url"))
        entry["thumbnail_url_preview"] = (m.get("thumbnail_url") or "")[:80]
        entry["has_image_versions2"] = bool(m.get("image_versions2"))
        entry["has_image_versions"] = bool(m.get("image_versions"))
        entry["has_carousel_media"] = bool(m.get("carousel_media"))
        entry["has_resources"] = bool(m.get("resources"))
        # carousel/resources 첫 아이템 구조
        for ckey in ("carousel_media", "resources"):
            cdata = m.get(ckey)
            if cdata and isinstance(cdata, list) and cdata:
                first = cdata[0]
                entry[f"{ckey}_first_keys"] = sorted(first.keys()) if isinstance(first, dict) else []
                if isinstance(first, dict):
                    entry[f"{ckey}_first_thumb"] = (first.get("thumbnail_url") or "")[:80]
                    entry[f"{ckey}_first_has_iv2"] = bool(first.get("image_versions2"))
                    entry[f"{ckey}_first_has_iv"] = bool(first.get("image_versions"))
        result.append(entry)
    return JSONResponse({"medias": result})


@app.get("/api/debug/db-posts/{pk}")
def debug_db_posts(pk: str):
    """DB posts 테이블에서 게시물 썸네일 상태 확인."""
    inf = get_influencer(pk)
    if not inf:
        return JSONResponse({"error": "인플루언서 없음"})
    posts = get_influencer_posts(pk)
    result = []
    for p in posts[:12]:
        result.append({
            "code": p.get("code", ""),
            "post_type": p.get("post_type", ""),
            "has_thumbnail": bool(p.get("thumbnail_url")),
            "thumbnail_preview": (p.get("thumbnail_url") or "")[:80],
            "likes": p.get("likes", 0),
            "views": p.get("views", 0),
        })
    return JSONResponse({"pk": pk, "username": inf.get("username"), "post_count": len(posts), "posts": result})


@app.get("/api/debug/test-crawl/{pk}")
def debug_test_crawl(pk: str, save: str = ""):
    """수집 테스트: 게시물 추출 + 선택적 DB 저장 (?save=1)."""
    inf = get_influencer(pk)
    if not inf:
        return JSONResponse({"error": "인플루언서 없음"})
    try:
        from crawler import _hiker_user_medias, _extract_media_fields, _enrich_reel_views
        from database import upsert_post
        medias = _hiker_user_medias(pk, amount=3)
        if not medias:
            return JSONResponse({"error": "HikerAPI 게시물 조회 실패"})
        _enrich_reel_views(medias)
        results = []
        save_errors = []
        for m in medias[:3]:
            try:
                post_data = _extract_media_fields(m, pk)
                results.append({
                    "post_id": post_data.get("post_id"),
                    "post_type": post_data.get("post_type"),
                    "thumbnail_url": (post_data.get("thumbnail_url") or "")[:80],
                    "likes": post_data.get("likes"),
                    "views": post_data.get("views"),
                    "code": post_data.get("post_url", ""),
                })
                if save == "1":
                    try:
                        upsert_post(post_data)
                    except Exception as se:
                        import traceback
                        save_errors.append({"post_id": post_data.get("post_id"), "error": str(se), "tb": traceback.format_exc()[-300:]})
            except Exception as e:
                import traceback
                results.append({"error": str(e), "tb": traceback.format_exc()[-300:]})
        resp = {"medias_count": len(medias), "results": results}
        if save == "1":
            resp["save_attempted"] = True
            resp["save_errors"] = save_errors
            # 저장 후 DB 확인
            posts = get_influencer_posts(pk)
            resp["db_post_count_after"] = len(posts)
        return JSONResponse(resp)
    except Exception as e:
        import traceback
        return JSONResponse({"error": str(e), "tb": traceback.format_exc()[-500:]})


@app.get("/api/debug/pdf-test/{pk}")
def debug_pdf_test(pk: str):
    """PDF 생성 테스트: 크기, 폰트, 버전 확인."""
    try:
        from export_pdf import export_single_pdf as _pdf, FONT_NAME
        inf, manual = _get_inf_with_manual(pk)
        data = _pdf(inf, manual)
        return JSONResponse({
            "ok": True,
            "pdf_bytes": len(data),
            "font": FONT_NAME,
            "username": inf.get("username"),
            "has_top_posts_likes": bool(inf.get("top_posts_likes")),
            "has_top_reels_views": bool(inf.get("top_reels_views")),
            "engagement_rate": inf.get("engagement_rate"),
        })
    except Exception as e:
        import traceback
        return JSONResponse({"error": str(e), "tb": traceback.format_exc()[-500:]})


@app.get("/api/debug/excluded-pks")
def debug_excluded_pks(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import get_banned_pks, get_hidden_pks
    banned = get_banned_pks()
    hidden = get_hidden_pks()
    return JSONResponse({
        "banned_count": len(banned), "banned_pks": list(banned)[:10],
        "hidden_count": len(hidden), "hidden_pks": list(hidden)[:10],
    })

@app.get("/api/collect-job/{job_id}")
def api_collect_job_status(job_id: int, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import get_collect_job
    job = get_collect_job(job_id)
    if not job: return JSONResponse({"error": "없음"}, 404)
    return JSONResponse(job)

@app.get("/api/collect-jobs")
def api_collect_jobs(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import get_collect_jobs
    jobs = get_collect_jobs(limit=20)
    # 시간 포맷
    from datetime import datetime, timezone, timedelta
    kst = timezone(timedelta(hours=9))
    for j in jobs:
        t = j.get("started_at")
        if not t:
            j["started_at_fmt"] = "-"
        elif isinstance(t, str) and "T" in t:
            try:
                dt = datetime.fromisoformat(t)
                j["started_at_fmt"] = dt.strftime("%m/%d %H:%M")
            except Exception:
                j["started_at_fmt"] = t[:16]
        else:
            try:
                j["started_at_fmt"] = datetime.fromtimestamp(float(t), tz=kst).strftime("%m/%d %H:%M")
            except Exception:
                j["started_at_fmt"] = "-"
        # finished_at 포맷
        ft = j.get("finished_at")
        if not ft:
            j["finished_at_fmt"] = "-"
        elif isinstance(ft, str) and "T" in ft:
            try:
                dft = datetime.fromisoformat(ft)
                j["finished_at_fmt"] = dft.strftime("%m/%d %H:%M")
            except Exception:
                j["finished_at_fmt"] = ft[:16]
        else:
            try:
                j["finished_at_fmt"] = datetime.fromtimestamp(float(ft), tz=kst).strftime("%m/%d %H:%M")
            except Exception:
                j["finished_at_fmt"] = "-"
    return JSONResponse(jobs)

@app.get("/api/collect-job/{job_id}/users")
def collect_job_users(job_id: int, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    from database import get_collect_job_users, get_collect_job
    job = get_collect_job(job_id)
    new_pks_set = set()
    if job:
        try:
            np = job.get("new_pks", "[]")
            new_pks_set = set(json.loads(np) if isinstance(np, str) else (np or []))
        except Exception:
            pass
    users = get_collect_job_users(job_id)
    for u in users:
        u["profile_pic_resolved"] = _pic(u)
        u["is_new"] = str(u.get("pk", "")) in new_pks_set
    return JSONResponse(users)


@app.post("/influencers/{pk}/refresh")
def refresh_one(pk: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    inf = get_influencer(pk)
    if not inf: return JSONResponse({"error": "없음"}, 404)

    from crawler import crawl_user_detail
    try:
        ok = crawl_user_detail(None, pk, inf["username"], inf.get("follower_count", 0))
        if ok:
            return JSONResponse({"ok": True})
        else:
            return JSONResponse({"ok": False, "error": "게시물을 가져올 수 없습니다 (API 응답 없음)"})
    except Exception as e:
        log.error(f"단일 갱신 실패 [{inf['username']}]: {e}", exc_info=True)
        return JSONResponse({"ok": False, "error": str(e)})


# ═══════════════════════════════════════════════════════
# 내보내기 (PDF / PPT / Excel)
# ═══════════════════════════════════════════════════════

def _get_inf_with_manual(pk):
    inf = get_influencer(pk)
    manual = get_manual(pk)
    for key in ["top_posts_likes", "top_posts_comments", "top_reels_views"]:
        try: inf[key] = json.loads(inf.get(key) or "[]")
        except: inf[key] = []
    return inf, manual


@app.get("/influencers/{pk}/export/pdf")
def export_single_pdf(pk: str, tpl: str = "scorecard",
                      session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id) or get_adv_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    try:
        from export_pdf import export_single_pdf as _pdf
        inf, manual = _get_inf_with_manual(pk)
        data = _pdf(inf, manual)
        tpl_label = {"scorecard": "스코어카드", "detail": "상세리포트"}.get(tpl, "스코어카드")
        fname = f"{inf.get('username', pk)}_{tpl_label}.pdf"
        return Response(data, media_type="application/pdf",
                        headers={"Content-Disposition": _safe_cd(fname)})
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        log.error(f"[PDF EXPORT ERROR] {e}\n{err}")
        _last_errors.append({"route": "pdf", "pk": pk, "error": str(e), "tb": err[-500:]})
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/influencers/{pk}/export/ppt")
def export_single_ppt(pk: str, tpl: str = "scorecard",
                      session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id) or get_adv_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    try:
        from export_ppt import export_single_ppt as _ppt
        inf, manual = _get_inf_with_manual(pk)
        data = _ppt(inf, manual)
        tpl_label = {"scorecard": "스코어카드", "proposal": "제안서"}.get(tpl, "스코어카드")
        fname = f"{inf.get('username', pk)}_{tpl_label}.pptx"
        return Response(data,
                        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                        headers={"Content-Disposition": _safe_cd(fname)})
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        log.error(f"[PPT EXPORT ERROR] {e}\n{err}")
        _last_errors.append({"route": "ppt", "pk": pk, "error": str(e), "tb": err[-500:]})
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/export/pdf")
def export_bulk_pdf(
    request: Request,
    mode: str = "scorecard",  # scorecard | list
    q: str = "", hashtag: str = "",
    min_f: str = "", max_f: str = "",
    verified: int = 0, public_only: int = 0,
    main_category: str = "", can_live: int = 0, only_approved: int = 0,
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id) or get_adv_user(session_id)
    if not user: return RedirectResponse("/login", 302)

    min_fi = int(min_f) if min_f and min_f.strip().isdigit() else None
    max_fi = int(max_f) if max_f and max_f.strip().isdigit() else None

    _, rows = get_influencers(
        keyword=q, hashtag_filter=hashtag, min_f=min_fi, max_f=max_fi,
        only_verified=bool(verified), exclude_private=bool(public_only),
        main_category=main_category, can_live=bool(can_live), only_approved=bool(only_approved),
        sort="follower_count", order="desc", page=1, per_page=500
    )
    pairs = [(_get_inf_with_manual(r["pk"])) for r in rows]

    if mode == "list":
        from export_pdf import export_list_pdf
        data = export_list_pdf(pairs)
        fname = f"influencer_list_{datetime.now().strftime('%Y%m%d')}.pdf"
    else:
        from export_pdf import export_multi_pdf
        data = export_multi_pdf(pairs)
        fname = f"influencer_scorecards_{datetime.now().strftime('%Y%m%d')}.pdf"

    return Response(data, media_type="application/pdf",
                    headers={"Content-Disposition": _safe_cd(fname)})


@app.get("/export/ppt")
def export_bulk_ppt(
    request: Request,
    mode: str = "scorecard",  # scorecard | list
    q: str = "", hashtag: str = "",
    min_f: str = "", max_f: str = "",
    verified: int = 0, public_only: int = 0,
    main_category: str = "", can_live: int = 0, only_approved: int = 0,
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id) or get_adv_user(session_id)
    if not user: return RedirectResponse("/login", 302)

    min_fi = int(min_f) if min_f and min_f.strip().isdigit() else None
    max_fi = int(max_f) if max_f and max_f.strip().isdigit() else None

    _, rows = get_influencers(
        keyword=q, hashtag_filter=hashtag, min_f=min_fi, max_f=max_fi,
        only_verified=bool(verified), exclude_private=bool(public_only),
        main_category=main_category, can_live=bool(can_live), only_approved=bool(only_approved),
        sort="follower_count", order="desc", page=1, per_page=300
    )
    pairs = [(_get_inf_with_manual(r["pk"])) for r in rows]

    if mode == "list":
        from export_ppt import export_list_ppt
        data = export_list_ppt(pairs)
    else:
        from export_ppt import export_multi_ppt
        data = export_multi_ppt(pairs)
    fname = f"influencer_{mode}_{datetime.now().strftime('%Y%m%d')}.pptx"
    return Response(data,
                    media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    headers={"Content-Disposition": _safe_cd(fname)})


@app.get("/export")
def export_excel(
    q: str = "", hashtag: str = "",
    min_f: Optional[str] = None, max_f: Optional[str] = None,
    verified: int = 0, public_only: int = 0,
    main_category: str = "", can_live: int = 0, only_approved: int = 0,
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id) or get_adv_user(session_id)
    if not user: return RedirectResponse("/login", 302)

    import openpyxl
    from io import BytesIO

    _min_f = int(min_f) if min_f and min_f.isdigit() else None
    _max_f = int(max_f) if max_f and max_f.isdigit() else None
    _, rows = get_influencers(
        keyword=q, hashtag_filter=hashtag, min_f=_min_f, max_f=_max_f,
        only_verified=bool(verified), exclude_private=bool(public_only),
        main_category=main_category, can_live=bool(can_live), only_approved=bool(only_approved),
        sort="follower_count", order="desc", page=1, per_page=100000
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "인플루언서"
    headers = ["username","full_name","팔로워","팔로우","게시물","인증","비공개",
               "카테고리","이메일","외부링크","해시태그",
               "참여율","평균좋아요","평균릴스뷰","릴스%","협찬%","업로드빈도",
               "마지막게시","라이브가능","피드단가","릴스단가","협업유형","품질점수","메모",
               "Instagram URL"]
    ws.append(headers)
    for r in rows:
        m = get_manual(r.get("pk", ""))
        ws.append([
            r.get("username"), r.get("full_name"),
            r.get("follower_count"), r.get("following_count"), r.get("media_count"),
            "O" if r.get("is_verified") else "",
            "O" if r.get("is_private") else "",
            r.get("category"), r.get("public_email"), r.get("external_url"), r.get("hashtags"),
            r.get("engagement_rate"), r.get("avg_likes"), r.get("avg_reel_views"),
            r.get("reels_ratio"), r.get("sponsored_ratio"), r.get("upload_frequency"),
            r.get("last_post_date"),
            "O" if m.get("can_live") else "",
            m.get("feed_price"), m.get("reel_price"),
            m.get("collab_types"), m.get("quality_score"), m.get("notes"),
            f"https://www.instagram.com/{r.get('username')}/",
        ])

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"influencers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _safe_cd(fname)})


# ═══════════════════════════════════════════════════════
# 엑셀 업로드 (외부 보고서 → DB 등록)
# ═══════════════════════════════════════════════════════

def _parse_kr_number(val) -> int:
    """한국어 숫자 파싱: '1,234', '7,000만', '1.2억' 등"""
    if val is None:
        return 0
    s = str(val).strip().replace(",", "")
    if "억" in s:
        try: return int(float(s.replace("억", "")) * 100_000_000)
        except: return 0
    if "만" in s:
        try: return int(float(s.replace("만", "")) * 10_000)
        except: return 0
    try: return int(float(s))
    except: return 0

def _yn_to_int(val) -> int:
    """Y/N → 1/0"""
    if val is None: return 0
    return 1 if str(val).strip().upper() in ("Y", "YES", "1", "TRUE") else 0

@app.get("/upload-excel", response_class=HTMLResponse)
def upload_excel_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = require_admin(session_id)
    return templates.TemplateResponse("upload_excel.html", {"request": request, "user": user})

@app.post("/api/upload-excel")
async def upload_excel_process(
    file: UploadFile = File(...),
    mode: str = Form("upsert"),
    hashtag_label: str = Form(""),
    session_id: Optional[str] = Cookie(default=None),
):
    user = require_admin(session_id)
    if not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse({"error": "xlsx 파일만 업로드 가능합니다."}, status_code=400)

    import openpyxl
    from io import BytesIO
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        return JSONResponse({"error": "파일 크기가 10MB를 초과합니다."}, status_code=400)

    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return JSONResponse({"error": f"엑셀 파일 읽기 실패: {e}"}, status_code=400)

    # 헤더 확인 (1행)
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    inf_rows = []
    man_rows = []
    skipped = 0
    errors = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            # AE열(index 30) = pk, A열(index 0) = username
            pk = str(row[30]).strip() if len(row) > 30 and row[30] else ""
            username = str(row[0]).strip() if row[0] else ""
            if not pk and not username:
                skipped += 1
                continue
            # pk가 없으면 username 기반으로 임시 pk 생성
            if not pk:
                pk = f"excel_{username}"

            follower = _parse_kr_number(row[2]) if len(row) > 2 else 0
            following = _parse_kr_number(row[3]) if len(row) > 3 else 0
            media = _parse_kr_number(row[6]) if len(row) > 6 else 0
            reel_count = _parse_kr_number(row[7]) if len(row) > 7 else 0
            avg_likes = _parse_kr_number(row[9]) if len(row) > 9 else 0
            avg_comments = _parse_kr_number(row[10]) if len(row) > 10 else 0
            is_private = int(bool(row[12])) if len(row) > 12 and row[12] else 0
            category = str(row[13]).strip() if len(row) > 13 and row[13] else ""
            is_business = _yn_to_int(row[14]) if len(row) > 14 else 0
            biz_category = str(row[15]).strip() if len(row) > 15 and row[15] else ""
            is_verified = _yn_to_int(row[18]) if len(row) > 18 else 0
            is_paid_verified = _yn_to_int(row[19]) if len(row) > 19 else 0
            location_info = str(row[20]).strip() if len(row) > 20 and row[20] else ""
            avg_reel_views = _parse_kr_number(row[21]) if len(row) > 21 else 0
            biography = str(row[23]).strip() if len(row) > 23 and row[23] else ""
            external_url = str(row[25]).strip() if len(row) > 25 and row[25] else ""
            highlight_count = _parse_kr_number(row[8]) if len(row) > 8 else 0
            comments_disabled = 1 if (len(row) > 11 and row[11] and str(row[11]).strip().upper() in ("Y", "YES", "1", "TRUE")) else 0
            is_professional = _yn_to_int(row[16]) if len(row) > 16 else 0
            has_threads = _yn_to_int(row[17]) if len(row) > 17 else 0

            # 카테고리: N열 우선, 없으면 P열(비지니스 카테고리)
            if not category and biz_category:
                category = biz_category

            # engagement_rate 계산
            engagement_rate = 0.0
            if follower > 0 and avg_likes > 0:
                engagement_rate = round((avg_likes + avg_comments) / follower * 100, 2)

            # feed_count 계산
            feed_count = max(0, media - reel_count)

            inf_data = {
                "pk": pk,
                "username": username,
                "full_name": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                "biography": biography,
                "follower_count": follower,
                "following_count": following,
                "media_count": media,
                "is_private": is_private,
                "is_verified": is_verified,
                "is_business": is_business,
                "category": category,
                "external_url": external_url,
                "avg_likes": avg_likes,
                "avg_comments": avg_comments,
                "engagement_rate": engagement_rate,
                "avg_reel_views": avg_reel_views,
                "reel_count": reel_count,
                "feed_count": feed_count,
                "highlight_count": highlight_count,
                "comments_disabled": comments_disabled,
                "has_threads": has_threads,
                "is_professional": is_professional,
                "is_paid_verified": is_paid_verified,
                "location_info": location_info,
            }
            # 해시태그 라벨이 있으면 추가
            if hashtag_label.strip():
                inf_data["hashtags"] = hashtag_label.strip()

            inf_rows.append(inf_data)

            # 매뉴얼 테이블 데이터 (페북, 스레드 등)
            facebook_url = str(row[27]).strip() if len(row) > 27 and row[27] else ""
            has_threads = _yn_to_int(row[17]) if len(row) > 17 else 0
            if facebook_url or has_threads:
                man_data = {"pk": pk}
                if facebook_url:
                    man_data["facebook_url"] = facebook_url
                if has_threads:
                    man_data["threads_url"] = "connected"
                man_rows.append(man_data)

        except Exception as e:
            errors += 1
            log.warning(f"엑셀 행 파싱 에러: {e}")
            continue

    wb.close()

    if not inf_rows:
        return JSONResponse({"error": "유효한 데이터가 없습니다.", "skipped": skipped, "errors": errors}, status_code=400)

    # DB에 벌크 upsert
    db_result = batch_upsert_from_excel(inf_rows, man_rows if man_rows else None)

    return JSONResponse({
        "ok": True,
        "total": len(inf_rows),
        "inserted": db_result.get("inserted", 0),
        "updated": db_result.get("updated", 0),
        "db_errors": db_result.get("errors", 0),
        "skipped": skipped,
        "parse_errors": errors,
    })


# ═══════════════════════════════════════════════════════
# 게시물 수집 (전체 갱신)
# ═══════════════════════════════════════════════════════

@app.get("/collect/posts", response_class=HTMLResponse)
def collect_posts_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    return templates.TemplateResponse("collect_posts.html", {
        "request": request, "user": user,
    })

@app.get("/target-extract", response_class=HTMLResponse)
def target_extract_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    return templates.TemplateResponse("target_extract.html", {"request": request, "user": user})


@app.get("/api/target-extract/stream")
def target_extract_stream(
    type: str = "followers", target: str = "",
    max_count: int = 500, save_to_db: str = "1",
    session_id: Optional[str] = Cookie(default=None),
):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    if not target: return JSONResponse({"error": "대상을 입력해주세요"}, 400)

    import requests as req
    token = os.getenv("HIKERAPI_TOKEN", "")
    save = save_to_db == "1"

    def stream():
        extracted = 0
        new_cnt = 0
        dup_cnt = 0
        cursor = None
        p = {"extracted": 0, "new_cnt": 0, "dup_cnt": 0, "max_count": max_count,
             "done": False, "users": [], "status": "", "label": ""}

        try:
            # 팔로워/팔로잉: username → user_id 변환
            if type in ("followers", "following"):
                label = "팔로워" if type == "followers" else "팔로잉"
                p["status"] = f"@{target} 계정 정보 조회 중"
                p["label"] = f"@{target} {label} 추출 중..."
                yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                # username → user_id
                info_r = req.get("https://api.hikerapi.com/v1/user/by/username",
                                 params={"username": target, "access_key": token}, timeout=15)
                info = info_r.json()
                user_id = str(info.get("pk", ""))
                if not user_id:
                    p.update({"done": True, "error": True, "status": f"@{target} 계정을 찾을 수 없습니다"})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                    return

                endpoint = f"https://api.hikerapi.com/v1/user/{type}/chunk"

                while extracted < max_count:
                    params = {"user_id": user_id, "access_key": token}
                    if cursor:
                        params["max_id"] = cursor

                    r = req.get(endpoint, params=params, timeout=15)
                    data = r.json()

                    if not isinstance(data, list) or len(data) < 2:
                        break
                    users_list = data[0] if isinstance(data[0], list) else []
                    cursor = data[1] if len(data) > 1 else None

                    if not users_list:
                        break

                    page_users = []
                    from database import upsert_influencer, get_influencer
                    for u in users_list:
                        if not isinstance(u, dict):
                            continue
                        if extracted >= max_count:
                            break
                        pk = str(u.get("pk", ""))
                        uname = u.get("username", "")
                        fname = u.get("full_name", "")
                        pic = u.get("profile_pic_url", "")
                        is_private = u.get("is_private", False)

                        existing = get_influencer(pk)
                        is_new = not existing
                        pic_local = ""

                        if save and not is_private:
                            try:
                                from database import upload_profile_pic
                                if pic:
                                    stored = upload_profile_pic(pk, pic)
                                    if stored: pic_local = stored
                                inf_data = {"pk": pk, "username": uname, "full_name": fname,
                                            "profile_pic_url": pic}
                                if pic_local: inf_data["profile_pic_local"] = pic_local
                                upsert_influencer(inf_data)
                            except Exception:
                                pass

                        if is_new: new_cnt += 1
                        else: dup_cnt += 1
                        extracted += 1

                        page_users.append({"username": uname, "full_name": fname,
                                          "pic": _pic({"profile_pic_local": pic_local, "profile_pic_url": pic, "username": uname}), "is_new": is_new})

                    p.update({"extracted": extracted, "new_cnt": new_cnt, "dup_cnt": dup_cnt,
                              "users": page_users,
                              "status": f"@{target} {label} 추출 중 — {extracted}명",
                              "label": f"@{target} {label} 추출 중..."})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                    if not cursor:
                        break
                    time.sleep(0.5)

            # 댓글 작성자
            elif type == "commenters":
                code = target.split("/p/")[-1].split("/")[0] if "/p/" in target else target.split("/reel/")[-1].split("/")[0] if "/reel/" in target else target
                p["status"] = f"게시물 정보 조회 중"
                p["label"] = "댓글 작성자 추출 중..."
                yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                # code → media_id
                media_r = req.get("https://api.hikerapi.com/v1/media/by/code",
                                  params={"code": code, "access_key": token}, timeout=15)
                media = media_r.json()
                media_id = str(media.get("pk", ""))
                if not media_id:
                    p.update({"done": True, "error": True, "status": "게시물을 찾을 수 없습니다"})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                    return

                endpoint = "https://api.hikerapi.com/v1/media/comments/chunk"
                seen_pks = set()

                while extracted < max_count:
                    params = {"id": media_id, "access_key": token}
                    if cursor:
                        params["max_id"] = cursor

                    r = req.get(endpoint, params=params, timeout=15)
                    data = r.json()

                    comments = []
                    next_cursor = None
                    if isinstance(data, list) and len(data) >= 2:
                        comments = data[0] if isinstance(data[0], list) else []
                        next_cursor = data[1] if len(data) > 1 else None
                    elif isinstance(data, dict):
                        comments = data.get("comments", [])
                        next_cursor = data.get("next_min_id")

                    if not comments:
                        break

                    page_users = []
                    from database import upsert_influencer, get_influencer
                    for c in comments:
                        if not isinstance(c, dict): continue
                        u = c.get("user", {})
                        if not isinstance(u, dict): continue
                        pk = str(u.get("pk", ""))
                        if not pk or pk in seen_pks: continue
                        seen_pks.add(pk)
                        if extracted >= max_count: break

                        uname = u.get("username", "")
                        fname = u.get("full_name", "")
                        pic = u.get("profile_pic_url", "")

                        existing = get_influencer(pk)
                        is_new = not existing

                        if save:
                            try:
                                upsert_influencer({"pk": pk, "username": uname,
                                                   "full_name": fname, "profile_pic_url": pic})
                            except Exception: pass

                        if is_new: new_cnt += 1
                        else: dup_cnt += 1
                        extracted += 1
                        page_users.append({"username": uname, "full_name": fname,
                                          "pic": _pic({"profile_pic_url": pic, "username": uname}), "is_new": is_new})

                    cursor = next_cursor
                    p.update({"extracted": extracted, "new_cnt": new_cnt, "dup_cnt": dup_cnt,
                              "users": page_users,
                              "status": f"댓글 작성자 추출 중 — {extracted}명"})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                    if not cursor: break
                    time.sleep(0.5)

            # 계정 전체 게시물 댓글 작성자
            elif type == "all_commenters":
                p["status"] = f"@{target} 계정 정보 조회 중"
                p["label"] = f"@{target} 전체 댓글 작성자 추출 중..."
                yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                info_r = req.get("https://api.hikerapi.com/v1/user/by/username",
                                 params={"username": target, "access_key": token}, timeout=15)
                info = info_r.json()
                user_id = str(info.get("pk", ""))
                if not user_id:
                    p.update({"done": True, "error": True, "status": f"@{target} 계정을 찾을 수 없습니다"})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                    return

                # 게시물 목록 가져오기
                medias_r = req.get("https://api.hikerapi.com/v1/user/medias/chunk",
                                   params={"user_id": user_id, "access_key": token}, timeout=15)
                medias_data = medias_r.json()
                media_items = []
                if isinstance(medias_data, list) and len(medias_data) >= 1:
                    media_items = medias_data[0] if isinstance(medias_data[0], list) else medias_data

                if not media_items:
                    p.update({"done": True, "error": True, "status": "게시물을 찾을 수 없습니다"})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                    return

                seen_pks = set()
                from database import upsert_influencer, get_influencer

                for mi, media in enumerate(media_items):
                    if extracted >= max_count:
                        break
                    if not isinstance(media, dict):
                        continue
                    media_id = str(media.get("pk", ""))
                    if not media_id:
                        continue

                    p["status"] = f"게시물 {mi+1:,}/{len(media_items):,} 댓글 추출 중 — {extracted:,}명"
                    p["users"] = []
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                    comment_cursor = None
                    while extracted < max_count:
                        cparams = {"id": media_id, "access_key": token}
                        if comment_cursor:
                            cparams["max_id"] = comment_cursor

                        cr = req.get("https://api.hikerapi.com/v1/media/comments/chunk",
                                     params=cparams, timeout=15)
                        cdata = cr.json()

                        comments = []
                        next_c = None
                        if isinstance(cdata, list) and len(cdata) >= 2:
                            comments = cdata[0] if isinstance(cdata[0], list) else []
                            next_c = cdata[1] if len(cdata) > 1 else None
                        elif isinstance(cdata, dict):
                            comments = cdata.get("comments", [])
                            next_c = cdata.get("next_min_id")

                        if not comments:
                            break

                        page_users = []
                        for c in comments:
                            if not isinstance(c, dict): continue
                            u = c.get("user", {})
                            if not isinstance(u, dict): continue
                            pk = str(u.get("pk", ""))
                            if not pk or pk in seen_pks: continue
                            seen_pks.add(pk)
                            if extracted >= max_count: break

                            uname = u.get("username", "")
                            fname = u.get("full_name", "")
                            pic = u.get("profile_pic_url", "")

                            existing = get_influencer(pk)
                            is_new = not existing

                            if save:
                                try:
                                    upsert_influencer({"pk": pk, "username": uname,
                                                       "full_name": fname, "profile_pic_url": pic})
                                except Exception: pass

                            if is_new: new_cnt += 1
                            else: dup_cnt += 1
                            extracted += 1
                            page_users.append({"username": uname, "full_name": fname,
                                              "pic": _pic({"profile_pic_url": pic, "username": uname}), "is_new": is_new})

                        comment_cursor = next_c
                        if page_users:
                            p.update({"extracted": extracted, "new_cnt": new_cnt, "dup_cnt": dup_cnt,
                                      "users": page_users,
                                      "status": f"게시물 {mi+1:,}/{len(media_items):,} — {extracted:,}명"})
                            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                        if not comment_cursor: break
                        time.sleep(0.3)

                    time.sleep(0.3)

            # 좋아요 누른 사람
            elif type == "likers":
                code = target.split("/p/")[-1].split("/")[0] if "/p/" in target else target.split("/reel/")[-1].split("/")[0] if "/reel/" in target else target
                p["status"] = "게시물 정보 조회 중"
                p["label"] = "좋아요 누른 사람 추출 중..."
                yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                media_r = req.get("https://api.hikerapi.com/v1/media/by/code",
                                  params={"code": code, "access_key": token}, timeout=15)
                media = media_r.json()
                media_id = str(media.get("pk", ""))
                if not media_id:
                    p.update({"done": True, "error": True, "status": "게시물을 찾을 수 없습니다"})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                    return

                likers_r = req.get("https://api.hikerapi.com/v1/media/likers",
                                   params={"id": media_id, "access_key": token}, timeout=15)
                likers_data = likers_r.json()
                users_list = likers_data if isinstance(likers_data, list) else likers_data.get("users", [])

                page_users = []
                from database import upsert_influencer, get_influencer
                for u in users_list:
                    if not isinstance(u, dict): continue
                    if extracted >= max_count: break
                    pk = str(u.get("pk", ""))
                    uname = u.get("username", "")
                    fname = u.get("full_name", "")
                    pic = u.get("profile_pic_url", "")

                    existing = get_influencer(pk)
                    is_new = not existing

                    if save:
                        try:
                            upsert_influencer({"pk": pk, "username": uname,
                                               "full_name": fname, "profile_pic_url": pic})
                        except Exception: pass

                    if is_new: new_cnt += 1
                    else: dup_cnt += 1
                    extracted += 1
                    page_users.append({"username": uname, "full_name": fname,
                                      "pic": _pic({"profile_pic_url": pic, "username": uname}), "is_new": is_new})

                p.update({"extracted": extracted, "new_cnt": new_cnt, "dup_cnt": dup_cnt,
                          "users": page_users,
                          "status": f"좋아요 추출 완료 — {extracted}명"})
                yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

            # 완료
            p.update({"done": True, "users": [],
                      "status": f"완료 — 추출 {extracted:,}명 (신규 {new_cnt:,} / 중복 {dup_cnt:,})"})
            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

        except Exception as e:
            log.error(f"타겟 추출 에러: {e}")
            p.update({"done": True, "error": True, "users": [],
                      "status": f"오류: {str(e)[:100]}"})
            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.get("/api/hashtag/search")
def hashtag_search_api(q: str = "", session_id: Optional[str] = Cookie(default=None)):
    """해시태그 검색 → 연관 해시태그 + 게시물 수"""
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    if not q: return JSONResponse([])
    import requests
    token = os.getenv("HIKERAPI_TOKEN", "")
    try:
        r = requests.get("https://api.hikerapi.com/v1/search/hashtags",
                         params={"query": q, "access_key": token}, timeout=10)
        data = r.json()
        if isinstance(data, list):
            return JSONResponse(data[:20])
    except Exception:
        pass
    return JSONResponse([])


@app.get("/api/location/search")
def api_location_search(q: str = Query(default=""), session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse([], 403)
    q = q.strip()
    if not q: return JSONResponse([])
    from crawler import _hiker_location_search
    results = _hiker_location_search(q)
    return JSONResponse(results)

@app.get("/refresh", response_class=HTMLResponse)
def refresh_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    cutoff = time.time() - 86400 * 30
    from database import get_banned_pks, get_hidden_pks
    banned_pks = get_banned_pks()
    hidden_pks = get_hidden_pks()
    excluded_pks = banned_pks | hidden_pks
    total_count, items = get_influencers(per_page=99999, page=1)
    stale = [r for r in items
             if (not r.get("stats_updated_at") or r["stats_updated_at"] < cutoff)
             and str(r.get("pk","")) not in excluded_pks]
    return templates.TemplateResponse("refresh.html", {
        "request": request, "user": user,
        "total_count": len(items), "stale_count": len(stale),
    })


@app.post("/refresh/start")
def refresh_start(session_id: Optional[str] = Cookie(default=None)):
    """게시물 수집 시작 → SSE 스트림으로 전환"""
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    return JSONResponse({"ok": True})

@app.get("/refresh/stream")
def refresh_stream(session_id: Optional[str] = Cookie(default=None)):
    """SSE 스트림 안에서 직접 게시물 갱신 실행 (서버리스 호환)"""
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)

    from crawler import crawl_user_detail

    def stream():
        try:
            cutoff = time.time() - 86400 * 30  # 30일 이상 지난 것만 갱신
            from database import get_banned_pks, get_hidden_pks
            banned_pks = get_banned_pks()
            hidden_pks = get_hidden_pks()
            excluded_pks = banned_pks | hidden_pks
            _, all_items = get_influencers(per_page=99999, page=1)
            rows = [r for r in all_items
                    if (not r.get("stats_updated_at") or r["stats_updated_at"] < cutoff)
                    and str(r.get("pk","")) not in excluded_pks]

            total = len(rows)
            success = fail = 0

            for i, row in enumerate(rows):
                pk = row.get("pk") or row.get("id")
                uname = row.get("username", "")
                followers = row.get("follower_count", 0)

                try:
                    ok = crawl_user_detail(None, str(pk), uname, followers or 0)
                    if ok:
                        success += 1
                    else:
                        fail += 1
                except Exception:
                    fail += 1

                p = {"running": True, "total": total, "done": i + 1,
                     "success": success, "fail": fail, "current_username": uname}
                yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                time.sleep(0.5)

            p = {"running": False, "total": total, "done": total,
                 "success": success, "fail": fail}
            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

        except Exception as e:
            log.error(f"게시물 갱신 에러: {e}")
            yield f"data: {json.dumps({'running': False, 'error': str(e)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

@app.post("/api/refresh-selected")
async def refresh_selected(request: Request, session_id: Optional[str] = Cookie(default=None)):
    """선택한 PK 목록만 갱신 (SSE 스트림)"""
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)

    body = await request.json()
    pks = body.get("pks", [])
    if not pks:
        return JSONResponse({"error": "선택된 항목이 없습니다."}, 400)

    from crawler import crawl_user_detail
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    def stream():
        try:
            total = len(pks)
            success = 0
            fail = 0
            done = 0
            lock = threading.Lock()

            yield f"data: {json.dumps({'running': True, 'total': total, 'done': 0, 'success': 0, 'fail': 0, 'current_username': '준비 중...'}, ensure_ascii=False)}\n\n"

            # PK별 정보 미리 조회
            tasks = []
            for pk in pks:
                inf = get_influencer(str(pk))
                uname = inf.get("username", "") if inf else str(pk)
                followers = inf.get("follower_count", 0) if inf else 0
                tasks.append((str(pk), uname, followers))

            # 3명씩 병렬 처리
            BATCH = 3
            for batch_start in range(0, len(tasks), BATCH):
                batch = tasks[batch_start:batch_start + BATCH]
                names = ", ".join(f"@{t[1]}" for t in batch)
                yield f"data: {json.dumps({'running': True, 'total': total, 'done': done, 'success': success, 'fail': fail, 'current_username': f'{names} 크롤링 중...'}, ensure_ascii=False)}\n\n"

                with ThreadPoolExecutor(max_workers=BATCH) as executor:
                    futures = {
                        executor.submit(crawl_user_detail, None, pk, uname, fol or 0): uname
                        for pk, uname, fol in batch
                    }
                    for future in as_completed(futures):
                        try:
                            ok = future.result()
                            if ok:
                                success += 1
                            else:
                                fail += 1
                        except Exception:
                            fail += 1
                        done += 1

                yield f"data: {json.dumps({'running': True, 'total': total, 'done': done, 'success': success, 'fail': fail, 'current_username': f'{done}/{total} 완료'}, ensure_ascii=False)}\n\n"

            yield f"data: {json.dumps({'running': False, 'total': total, 'done': total, 'success': success, 'fail': fail}, ensure_ascii=False)}\n\n"
        except Exception as e:
            log.error(f"선택 갱신 에러: {e}")
            yield f"data: {json.dumps({'running': False, 'error': str(e)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

@app.get("/refresh/status")
def refresh_status(session_id: Optional[str] = Cookie(default=None)):
    """하위호환용 — SSE 방식으로 전환됨"""
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    return JSONResponse({})


# ═══════════════════════════════════════════════════════
# 자동화 (Cron)
# ═══════════════════════════════════════════════════════

@app.get("/api/cron/auto")
def cron_auto(request: Request):
    """Vercel Cron 또는 외부 cron이 호출하는 자동화 엔드포인트.
    1) auto_collect 해시태그 1개 → 신규 계정 10명 수집
    2) 미갱신 인플루언서 5명 → 게시물 갱신
    """
    # 인증: Vercel CRON_SECRET 또는 커스텀 헤더
    cron_secret = os.getenv("CRON_SECRET", "")
    auth = request.headers.get("authorization", "")
    if cron_secret and auth != f"Bearer {cron_secret}":
        raise HTTPException(401, "Unauthorized")
    if not cron_secret:
        raise HTTPException(403, "CRON_SECRET 미설정")

    from crawler import cron_collect_batch, cron_refresh_batch

    results = {"collect": None, "refresh": None}

    # ① 계정 수집: auto_collect 해시태그 중 가장 오래된 것
    auto_tags = get_auto_hashtags()
    if auto_tags:
        tag = auto_tags[0]
        tag_name = tag.get("name", "")
        try:
            r = cron_collect_batch(tag_name, target_users=0, search_type="recent")
            results["collect"] = {"hashtag": tag_name, **r}
            add_cron_log("collect", "error" if r.get("error") else "ok",
                         hashtag=tag_name, details=r)
        except Exception as e:
            results["collect"] = {"hashtag": tag_name, "error": str(e)}
            add_cron_log("collect", "error", hashtag=tag_name, details={"error": str(e)})

    # ② 게시물 갱신
    try:
        r = cron_refresh_batch(batch_size=5, stale_hours=720)
        results["refresh"] = r
        add_cron_log("refresh", "ok", details=r)
    except Exception as e:
        results["refresh"] = {"error": str(e)}
        add_cron_log("refresh", "error", details={"error": str(e)})

    return JSONResponse(results)


@app.get("/api/hiker-balance")
def hiker_balance(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    token = os.getenv("HIKERAPI_TOKEN", "").strip()
    if not token:
        return JSONResponse({"error": "HIKERAPI_TOKEN 미설정"}, 400)
    try:
        import requests as _r
        r = _r.get("https://api.hikerapi.com/sys/balance",
                    headers={"x-access-key": token}, timeout=10)
        return JSONResponse(r.json())
    except Exception as e:
        return JSONResponse({"error": str(e)}, 500)


@app.post("/api/cron/manual-run")
def cron_manual_run(session_id: Optional[str] = Cookie(default=None)):
    """관리자가 대시보드에서 수동으로 cron 1회 실행"""
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)

    from crawler import cron_collect_batch, cron_refresh_batch

    results = {"collect": None, "refresh": None}

    auto_tags = get_auto_hashtags()
    if auto_tags:
        tag = auto_tags[0]
        tag_name = tag.get("name", "")
        try:
            r = cron_collect_batch(tag_name, target_users=0, search_type="recent")
            results["collect"] = {"hashtag": tag_name, **r}
            add_cron_log("collect", "error" if r.get("error") else "ok",
                         hashtag=tag_name, details=r)
        except Exception as e:
            results["collect"] = {"hashtag": tag_name, "error": str(e)}
            add_cron_log("collect", "error", hashtag=tag_name, details={"error": str(e)})

    try:
        r = cron_refresh_batch(batch_size=5, stale_hours=720)
        results["refresh"] = r
        add_cron_log("refresh", "ok", details=r)
    except Exception as e:
        results["refresh"] = {"error": str(e)}
        add_cron_log("refresh", "error", details={"error": str(e)})

    return JSONResponse(results)


@app.get("/automation", response_class=HTMLResponse)
def automation_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    stats = get_stats()
    auto_tags = get_auto_hashtags()
    cron_logs = get_cron_logs(30)
    return templates.TemplateResponse("automation.html", {
        "request": request, "user": user,
        "stats": stats, "auto_tags": auto_tags, "cron_logs": cron_logs,
        "cron_secret_set": bool(os.getenv("CRON_SECRET", "")),
    })


# ═══════════════════════════════════════════════════════
# 해시태그 관리
# ═══════════════════════════════════════════════════════

@app.get("/hashtags", response_class=HTMLResponse)
def hashtags_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    rows = get_hashtags()
    jobs = get_collect_jobs(30)
    return templates.TemplateResponse("hashtags.html", {
        "request": request, "user": user,
        "hashtags": [dict(r) if not isinstance(r, dict) else r for r in rows],
        "jobs": [dict(r) if not isinstance(r, dict) else r for r in jobs],
    })

@app.post("/hashtags/add")
def add_hashtag_route(name: str = Form(...), requested_count: int = Form(default=500),
                auto_collect: int = Form(default=1),
                session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    name = name.strip().lstrip("#")
    try:
        db_add_hashtag(name, requested_count, auto_collect)
    except Exception as e:
        log.error(f"해시태그 추가 실패: {e}")
    return RedirectResponse("/hashtags", 302)

@app.post("/hashtags/delete")
def delete_hashtag_route(name: str = Form(default=""), hashtag_id: int = Form(default=0),
                         session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    if hashtag_id:
        db_delete_hashtag(hashtag_id)
    elif name:
        # SQLite fallback: find by name
        from database import _USE_SUPABASE
        if not _USE_SUPABASE:
            conn = get_conn()
            conn.execute("DELETE FROM insta_hashtags WHERE name=?", (name,))
            conn.commit(); conn.close()
    return RedirectResponse("/hashtags", 302)


# ═══════════════════════════════════════════════════════
# 수집
# ═══════════════════════════════════════════════════════

@app.get("/collect", response_class=HTMLResponse)
def collect_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    from database import get_collect_jobs
    htags = get_hashtags()
    jobs = get_collect_jobs(limit=20)
    return templates.TemplateResponse("collect.html", {
        "request": request, "user": user,
        "hashtags": [r.get("name","") if isinstance(r, dict) else r["name"] for r in htags],
        "jobs": jobs,
    })

@app.post("/collect/start")
def collect_start(hashtag: str = Form(default=""), requested_count: int = Form(default=500),
                  target_users: int = Form(default=0),
                  search_type: str = Form(default="recent"),
                  collect_mode: str = Form(default="hashtag"),
                  location_pk: str = Form(default=""),
                  location_name: str = Form(default=""),
                  session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return HTMLResponse("인증 필요", 403)
    from database import add_collect_job
    if collect_mode == "location" and location_pk:
        label = location_name or f"위치:{location_pk}"
        job_db_id = add_collect_job(label, "running", target_users, search_type)
        return HTMLResponse(f"{job_db_id}|location|{location_pk}")
    else:
        hashtag = hashtag.strip().lstrip("#")
        job_db_id = add_collect_job(hashtag, "running", target_users, search_type)
        return HTMLResponse(str(job_db_id))

@app.post("/collect/stop/{job_id}")
async def collect_stop(request: Request, job_id: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return JSONResponse({"error": "인증 필요"}, 403)
    try:
        jid = int(job_id)
        # 클라이언트에서 최종 수치 전달 시 함께 저장
        save_data = {"status": "stopped"}
        try:
            body = await request.json()
            if body.get("collected_posts") is not None:
                save_data["collected_posts"] = int(body["collected_posts"])
            if body.get("new_users") is not None:
                save_data["new_users"] = int(body["new_users"])
            if body.get("updated_users") is not None:
                save_data["updated_users"] = int(body["updated_users"])
            if body.get("last_next_id"):
                save_data["last_next_id"] = str(body["last_next_id"])
            if body.get("last_page") is not None:
                save_data["last_page"] = int(body["last_page"])
        except Exception:
            pass
        from datetime import datetime, timezone, timedelta
        save_data["finished_at"] = datetime.now(timezone(timedelta(hours=9))).isoformat()
        update_collect_job(jid, **save_data)
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"error": str(e)}, 500)

@app.get("/collect/progress/{job_id}")
def collect_progress(job_id: str,
                     hashtag: str = Query(default=""),
                     target_users: int = Query(default=30),
                     search_type: str = Query(default="recent"),
                     collect_mode: str = Query(default="hashtag"),
                     location_pk: str = Query(default=""),
                     resume_from: str = Query(default=""),
                     resume_new: int = Query(default=0),
                     resume_updated: int = Query(default=0),
                     resume_posts: int = Query(default=0),
                     resume_page: int = Query(default=0),
                     session_id: Optional[str] = Cookie(default=None)):
    """SSE 스트림 안에서 직접 수집 실행 (서버리스 호환). resume_from으로 이어서 수집 가능."""
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)

    from database import upsert_influencer, update_collect_job, get_existing_pks, get_banned_pks

    job_db_id = None
    try:
        job_db_id = int(job_id)
    except (ValueError, TypeError):
        return JSONResponse({"error": "잘못된 job_id"}, 400)

    def stream():
        _is_location = collect_mode == "location" and location_pk
        _label = hashtag or (f"위치:{location_pk}" if _is_location else "")
        if not hashtag and not _is_location:
            yield f"data: {json.dumps({'done': True, 'error': '해시태그 또는 위치를 입력하세요'}, ensure_ascii=False)}\n\n"
            return

        if not _is_location:
            try:
                update_hashtag_status(hashtag, "running")
            except Exception:
                try: db_add_hashtag(hashtag)
                except: pass

        # 기존 job 상태를 running으로 업데이트
        try:
            update_collect_job(job_db_id, status="running")
        except Exception:
            pass

        log.info(f"[{_label}] SSE 시작 — resume_from={resume_from}, resume_new={resume_new}, resume_posts={resume_posts}, resume_page={resume_page}")

        BATCH_PAGES = 500  # 재연결 없이 한번에 끝까지 수집 (기존 15 → 500)

        try:
            from crawler import _hiker_hashtag_medias_page, _hiker_location_medias_page

            existing_pks = get_existing_pks()
            banned_pks = get_banned_pks()

            seen_pks = set()
            _seen_post_pks = set()  # 게시물 PK 중복 추적
            collected_pk_list = []
            new_pk_list = []
            # DB에 저장된 최신 값과 클라이언트 값 비교 → 큰 값 사용
            total_medias = resume_posts
            new_cnt = resume_new
            updated_cnt = resume_updated
            max_id = resume_from or None
            _resume_page = resume_page
            try:
                from database import get_collect_job as _gcj_init
                _db_job = _gcj_init(job_db_id)
                if _db_job:
                    total_medias = max(resume_posts, int(_db_job.get("collected_posts", 0) or 0))
                    new_cnt = max(resume_new, int(_db_job.get("new_users", 0) or 0))
                    updated_cnt = max(resume_updated, int(_db_job.get("updated_users", 0) or 0))
                    _db_next = str(_db_job.get("last_next_id", "") or "")
                    _db_page = int(_db_job.get("last_page", 0) or 0)
                    if _db_page > _resume_page:
                        _resume_page = _db_page
                    if _db_next and not max_id:
                        max_id = _db_next
            except Exception:
                pass
            # search_type에 따라 메인 엔드포인트 결정 (각각 독립 작동)
            # recent = v1/hashtag/medias/top/recent/chunk (최근 게시물)
            # top    = v1/hashtag/medias/top/chunk (인기 게시물)
            if search_type == "recent":
                _ENDPOINTS = ["recent", "clips"]
            else:
                _ENDPOINTS = ["top", "clips"]
            # 재연결 시 이전 엔드포인트 복원 (last_next_id에 "ep:idx:" 접두어로 저장)
            _ep_idx = 0
            if max_id and max_id.startswith("ep:"):
                try:
                    parts = max_id.split(":", 3)  # "ep:idx:real_cursor"
                    _ep_idx = int(parts[1])
                    max_id = parts[2] if len(parts) > 2 and parts[2] else None
                    if _ep_idx >= len(_ENDPOINTS):
                        _ep_idx = 0
                except Exception:
                    _ep_idx = 0
            endpoint = _ENDPOINTS[_ep_idx]
            page_num = _resume_page
            batch_done = 0
            last_next_id = None
            no_data = False

            _consec_dup_pages = 0  # 연속 100% 중복 게시물 페이지 수 (순환 감지)

            # DB 보정된 값으로 첫 SSE 메시지 전송
            p = {"hashtag": _label, "posts": total_medias, "new": new_cnt,
                 "updated": updated_cnt, "done": False, "error": None,
                 "status": "게시물 검색 중", "target": target_users}
            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

            while new_cnt < target_users and batch_done < BATCH_PAGES:
                # 다른 탭에서 중지 요청 확인 (3페이지마다 — DB 호출 절약)
                if batch_done % 3 == 0:
                    try:
                        from database import get_collect_job as _gcj2
                        _job_check = _gcj2(job_db_id)
                        if _job_check and _job_check.get("status") == "stopped":
                            try:
                                from datetime import datetime as _dt3, timezone as _tz3, timedelta as _td3
                                _stop_next = f"ep:{_ep_idx}:{last_next_id}" if last_next_id else ""
                                update_collect_job(job_db_id,
                                    collected_posts=total_medias, new_users=new_cnt,
                                    updated_users=updated_cnt, last_next_id=_stop_next,
                                    last_page=page_num, status="stopped",
                                    finished_at=_dt3.now(_tz3(_td3(hours=9))).isoformat())
                            except Exception:
                                pass
                            p.update({"done": True, "status": f"중지됨 — 신규 {new_cnt:,}명 저장됨",
                                      "new": new_cnt, "updated": updated_cnt, "posts": total_medias})
                            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                            return
                    except Exception as _stop_err:
                        log.warning(f"Stop check failed: {_stop_err}")

                page_num += 1
                batch_done += 1
                _page_start = time.time()

                def _fetch_page(mid):
                    if _is_location:
                        return _hiker_location_medias_page(location_pk, endpoint, mid)
                    else:
                        return _hiker_hashtag_medias_page(hashtag, endpoint, mid)

                try:
                    items, next_id = _fetch_page(max_id)
                except Exception as api_err:
                    # API 에러 시 1회 재시도
                    time.sleep(2)
                    try:
                        items, next_id = _fetch_page(max_id)
                    except Exception:
                        raise api_err

                if not items and not max_id:
                    _src = "위치" if _is_location else "해시태그"
                    raise Exception(f"HikerAPI {_src} 조회 실패 — HIKERAPI_TOKEN을 확인하세요")
                if not items and max_id:
                    # 커서가 만료됐을 수 있음 → 처음부터 재시도
                    time.sleep(1)
                    items, next_id = _fetch_page(max_id)
                    if not items and resume_from:
                        # resume 커서 만료 → 처음부터 시작
                        max_id = None
                        p.update({"status": "커서 만료 — 처음부터 다시 수집"})
                        yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                        time.sleep(0.5)
                        items, next_id = _fetch_page(None)
                    if not items:
                        no_data = True
                        break

                total_medias += len(items)
                page_users = []
                _new_batch = []  # (pk_str, uname, fname, pic, hashtag) 배치 삽입용
                page_post_pks = []  # 이 페이지의 게시물 PK 목록 (UI 표시용)
                page_dup_post = 0  # 이 페이지의 중복 게시물 수

                for m in items:
                    if not isinstance(m, dict):
                        continue
                    # 게시물 PK 추적
                    post_pk = str(m.get("pk") or m.get("id") or "")
                    if post_pk:
                        if post_pk in _seen_post_pks:
                            page_dup_post += 1
                        else:
                            _seen_post_pks.add(post_pk)
                        page_post_pks.append(post_pk)

                    user_data = m.get("user") or {}
                    if not isinstance(user_data, dict):
                        continue
                    upk = user_data.get("pk")
                    if not upk:
                        continue
                    pk_str = str(upk)
                    if pk_str in seen_pks:
                        continue
                    seen_pks.add(pk_str)
                    if pk_str in banned_pks:
                        continue
                    uname = user_data.get("username", "")
                    fname = user_data.get("full_name", "")
                    pic = str(user_data.get("profile_pic_url", "") or "")
                    is_new = pk_str not in existing_pks

                    if is_new:
                        _new_batch.append((pk_str, uname, fname, pic))
                        new_cnt += 1
                        new_pk_list.append(pk_str)
                        existing_pks.add(pk_str)
                    else:
                        updated_cnt += 1
                    collected_pk_list.append(pk_str)

                    page_users.append({
                        "username": uname, "full_name": fname,
                        "pic": _pic({"profile_pic_url": pic, "username": uname}),
                        "is_new": is_new,
                    })

                # ── 신규 유저 배치 DB 삽입 (프로필 사진 업로드는 나중에) ──
                if _new_batch:
                    from database import batch_insert_influencers
                    _tag = hashtag or _label
                    batch_insert_influencers([
                        {"pk": nb[0], "username": nb[1], "full_name": nb[2],
                         "profile_pic_url": nb[3], "hashtag": _tag}
                        for nb in _new_batch
                    ])

                # 클라이언트에 ep 접두어 포함된 next_id 전달 (재연결 시 엔드포인트 복원용)
                _client_next = f"ep:{_ep_idx}:{next_id}" if next_id else ""
                p.update({"posts": total_medias, "new": new_cnt, "updated": updated_cnt,
                          "unique_posts": len(_seen_post_pks), "dup_posts": page_dup_post,
                          "post_pks": page_post_pks[-5:],  # 최근 5개 게시물 PK
                          "status": f"p{page_num:,} [{endpoint}] — 신규 {new_cnt:,}명 / 중복 {updated_cnt:,}명 / 고유게시물 {len(_seen_post_pks):,}개 (이번페이지 중복 {page_dup_post}개)",
                          "page": page_num, "page_items": len(items),
                          "has_next": bool(next_id),
                          "next_id": _client_next,
                          "users": page_users})
                yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

                # 3페이지마다 DB 통계 업데이트 (매 페이지 → 불필요한 오버헤드)
                if batch_done % 3 == 0 or not next_id:
                    # last_next_id에 엔드포인트 인덱스를 접두어로 저장 (재연결 시 복원용)
                    _save_next = f"ep:{_ep_idx}:{last_next_id}" if last_next_id else ""
                    try:
                        update_collect_job(job_db_id,
                            collected_posts=total_medias, new_users=new_cnt,
                            updated_users=updated_cnt, status="running",
                            last_next_id=_save_next, last_page=page_num)
                    except Exception:
                        pass

                # 순환 감지: 이 페이지 게시물이 전부 이미 본 것이면 카운트
                if page_dup_post >= len(items) * 0.9:  # 90% 이상 중복
                    _consec_dup_pages += 1
                else:
                    _consec_dup_pages = 0

                # 3페이지 연속 90%+ 중복 → 현재 엔드포인트 순환으로 판단, 전환
                if _consec_dup_pages >= 3:
                    log.info(f"[{_label}] {endpoint} 순환 감지 ({_consec_dup_pages}페이지 연속 중복) → 전환")
                    p.update({"status": f"{endpoint} 순환 감지 — 다음 엔드포인트로 전환"})
                    yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                    next_id = None  # 강제로 엔드포인트 전환 트리거
                    _consec_dup_pages = 0

                if not next_id:
                    # 현재 엔드포인트 소진 → 다음 엔드포인트로 전환
                    _ep_idx += 1
                    if _ep_idx < len(_ENDPOINTS):
                        _prev_ep = endpoint
                        endpoint = _ENDPOINTS[_ep_idx]
                        max_id = None
                        # 엔드포인트 전환 시 즉시 DB에 저장 (재연결 대비)
                        try:
                            update_collect_job(job_db_id,
                                last_next_id=f"ep:{_ep_idx}:", last_page=page_num)
                        except Exception:
                            pass
                        log.info(f"[{_label}] {_prev_ep} 소진 → {endpoint} 전환 (신규 {new_cnt})")
                        p.update({"status": f"{endpoint} 게시물 검색 중 — 신규 {new_cnt:,}명"})
                        yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"
                        continue
                    no_data = True
                    break
                max_id = next_id
                last_next_id = next_id

                # 적응형 대기: API 응답이 느리면 대기 줄임
                elapsed = time.time() - _page_start
                if elapsed > 3:
                    time.sleep(0.1)  # API가 이미 느렸으면 최소 대기
                else:
                    time.sleep(max(0.2, 0.5 - elapsed * 0.1))

            # 목표 달성 or 더 이상 데이터 없음 → 진짜 완료
            is_truly_done = new_cnt >= target_users or no_data

            # 배치 결과 저장 (진행중이면 running 유지, 완료시만 done)
            try:
                # 기존 collected_pks에 누적
                from database import get_collect_job
                prev_job = get_collect_job(job_db_id)
                prev_pks = []
                prev_new_pks = []
                if prev_job:
                    try:
                        pp = prev_job.get("collected_pks", "[]")
                        prev_pks = json.loads(pp) if isinstance(pp, str) else (pp or [])
                    except: pass
                    try:
                        pn = prev_job.get("new_pks", "[]")
                        prev_new_pks = json.loads(pn) if isinstance(pn, str) else (pn or [])
                    except: pass
                all_pks = list(set(prev_pks + collected_pk_list))
                all_new_pks = list(set(prev_new_pks + new_pk_list))

                _save_next_final = f"ep:{_ep_idx}:{last_next_id}" if last_next_id else ""
                save_data = dict(
                    collected_posts=total_medias, new_users=new_cnt,
                    updated_users=updated_cnt,
                    collected_pks=json.dumps(all_pks),
                    new_pks=json.dumps(all_new_pks),
                    last_next_id=_save_next_final,
                    last_page=page_num)
                if is_truly_done:
                    save_data["status"] = "done"
                    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
                    save_data["finished_at"] = _dt.now(_tz(_td(hours=9))).isoformat()
                else:
                    save_data["status"] = "running"
                update_collect_job(job_db_id, **save_data)
                if is_truly_done:
                    update_hashtag_status(hashtag, "idle")
            except Exception:
                pass

            if is_truly_done:
                reason = "목표 달성" if new_cnt >= target_users else f"해시태그 끝 ({page_num}페이지)"
                p.update({"done": True, "new": new_cnt, "updated": updated_cnt,
                          "status": f"완료 — 신규 {new_cnt:,}명 / 중복 {updated_cnt:,}명 ({reason})",
                          "page": page_num})
            else:
                _resume_next = f"ep:{_ep_idx}:{last_next_id}" if last_next_id else ""
                p.update({"done": False, "has_more": True, "new": new_cnt, "updated": updated_cnt,
                          "next_id": _resume_next, "page": page_num,
                          "posts": total_medias,
                          "status": f"배치 {page_num:,}페이지 완료 — 신규 {new_cnt:,}명 / 중복 {updated_cnt:,}명 / 목표 {target_users:,}명 (자동 계속)"})
            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

        except Exception as e:
            log.error(f"[{hashtag}] 수집 에러: {e}")
            try:
                if job_db_id:
                    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
                    _now_iso = _dt.now(_tz(_td(hours=9))).isoformat()
                    update_collect_job(job_db_id, status="error", error_msg=str(e)[:200], finished_at=_now_iso)
                update_hashtag_status(hashtag, "error")
            except Exception:
                pass
            p.update({"done": True, "error": str(e), "status": "수집 실패"})
            yield f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ═══════════════════════════════════════════════════════
# 계정 수집 (Xpoz API)
# ═══════════════════════════════════════════════════════

_XPOZ_API_KEY = os.getenv("XPOZ_API_KEY", "")

@app.get("/xpoz-collect", response_class=HTMLResponse)
def xpoz_collect_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return RedirectResponse("/login", 302)
    return templates.TemplateResponse("xpoz_collect.html", {"request": request, "user": user})

@app.get("/api/xpoz-collect/stream")
def xpoz_collect_stream(
    keyword: str = Query(default=""),
    target_users: int = Query(default=1000),
    search_mode: str = Query(default="posts"),
    sort_mode: str = Query(default="relevant"),
    start_date: str = Query(default=""),
    end_date: str = Query(default=""),
    session_id: Optional[str] = Cookie(default=None),
):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    if not keyword:
        return JSONResponse({"error": "키워드 필수"}, 400)

    api_key = _XPOZ_API_KEY
    if not api_key:
        return JSONResponse({"error": "XPOZ_API_KEY 환경변수 필요"}, 400)

    from database import upsert_influencer, get_existing_pks, get_banned_pks

    def stream():
        from xpoz import XpozClient
        from datetime import datetime, timedelta

        client = XpozClient(api_key)
        ig = client.instagram

        existing_pks = set()
        try:
            existing_pks = set(get_existing_pks())
        except Exception:
            pass
        banned_pks = set()
        try:
            banned_pks = set(get_banned_pks())
        except Exception:
            pass

        keywords = [k.strip() for k in keyword.split(",") if k.strip()]
        new_count = 0
        dup_count = 0
        total_pages = 0
        total_results = 0
        credits_used = 0
        seen_usernames = set()

        _start = start_date if start_date else None
        _end = end_date if end_date else None

        def _make_progress(kw_idx, kw, kw_new, kw_dup, kw_page, status, batch_users=None, done=False, error=None):
            p = {
                "new": new_count, "dup": dup_count,
                "pages": total_pages, "results": total_results,
                "credits": round(credits_used, 1),
                "current_keyword": kw, "kw_new": kw_new, "kw_dup": kw_dup, "kw_page": kw_page,
                "status": status,
            }
            if batch_users:
                p["users"] = batch_users
            if done:
                p["done"] = True
            if error:
                p["error"] = error
            return f"data: {json.dumps(p, ensure_ascii=False)}\n\n"

        def _extract_user_id(item, mode):
            """유저 ID 추출: users 모드는 id가 유저 PK, posts 모드는 id에서 파싱"""
            if mode == "users":
                return str(getattr(item, 'id', '') or '')
            # posts 모드: id 형식 = "{post_id}_{user_id}"
            raw_id = str(getattr(item, 'id', '') or '')
            if '_' in raw_id:
                return raw_id.split('_')[-1]
            return ''

        yield _make_progress(0, '', 0, 0, 0, f'{len(keywords)}개 키워드 수집 시작...')

        for kw_idx, kw in enumerate(keywords):
            if new_count >= target_users:
                break

            kw_new = 0
            kw_dup = 0
            kw_page = 0

            try:
                kwargs = {"query": kw}
                if _start:
                    kwargs["start_date"] = _start
                if _end:
                    kwargs["end_date"] = _end

                if search_mode == "users":
                    result = ig.get_users_by_keywords(**kwargs)
                else:
                    if sort_mode == "latest":
                        kwargs["force_latest"] = True
                    result = ig.search_posts(**kwargs)

                credits_used += 5  # 첫 쿼리 5크레딧

                while True:
                    if not result or not hasattr(result, 'data') or not result.data:
                        break

                    kw_page += 1
                    total_pages += 1
                    batch_users = []
                    page_new = 0

                    for item in result.data:
                        username = getattr(item, 'username', None) or ''
                        if not username or username in seen_usernames:
                            continue
                        seen_usernames.add(username)

                        user_id = _extract_user_id(item, search_mode)
                        full_name = getattr(item, 'full_name', '') or ''

                        is_new = user_id not in existing_pks if user_id else True
                        if user_id and user_id in banned_pks:
                            continue

                        if is_new:
                            new_count += 1
                            kw_new += 1
                            page_new += 1
                            if user_id:
                                try:
                                    upsert_influencer({
                                        "pk": user_id,
                                        "username": username,
                                        "full_name": full_name,
                                        "profile_pic_url": getattr(item, 'profile_pic_url', '') or '',
                                        "source_hashtag": kw,
                                    })
                                    existing_pks.add(user_id)
                                except Exception as db_err:
                                    log.warning(f"[xpoz] DB upsert 에러: {db_err}")
                        else:
                            dup_count += 1
                            kw_dup += 1

                        batch_users.append({
                            "username": username,
                            "full_name": full_name,
                            "is_new": is_new,
                        })

                    total_results += len(result.data)
                    credits_used += max(1, len(result.data)) * 0.005

                    status_msg = f"[{kw_idx+1}/{len(keywords)}] \"{kw}\" p{kw_page} — 신규 {new_count}명 / 중복 {dup_count}명 (이 페이지 +{page_new}명)"

                    if new_count >= target_users:
                        yield _make_progress(kw_idx, kw, kw_new, kw_dup, kw_page,
                            f"목표 달성! 신규 {new_count}명 수집 완료 ({total_results}건 검색)",
                            batch_users, done=True)
                        client.close()
                        return

                    yield _make_progress(kw_idx, kw, kw_new, kw_dup, kw_page, status_msg, batch_users)

                    # 신규 0명이면 즉시 다음 키워드로 (크레딧 절약)
                    if page_new == 0 and kw_page >= 2:
                        log.info(f"[xpoz] \"{kw}\" p{kw_page} 신규 0명 → 다음 키워드")
                        yield _make_progress(kw_idx, kw, kw_new, kw_dup, kw_page,
                            f"\"{kw}\" 신규 없음 → 다음 키워드로 (p{kw_page})")
                        break

                    # 다음 페이지 (SDK next_page()는 total_pages=0 버그, 직접 호출)
                    try:
                        next_result = result._fetch_page_result(kw_page + 1)
                        if next_result is None or not hasattr(next_result, 'data') or not next_result.data:
                            log.info(f"[xpoz] \"{kw}\" 더 이상 결과 없음 (p{kw_page})")
                            yield _make_progress(kw_idx, kw, kw_new, kw_dup, kw_page,
                                f"\"{kw}\" 마지막 페이지 (p{kw_page})")
                            break
                        result = next_result
                        credits_used += 5
                    except Exception as page_err:
                        log.warning(f"[xpoz] \"{kw}\" 페이지 에러 (p{kw_page}): {page_err}")
                        yield _make_progress(kw_idx, kw, kw_new, kw_dup, kw_page,
                            f"\"{kw}\" 마지막 페이지 (p{kw_page})")
                        break

            except Exception as e:
                err_msg = str(e)[:150]
                log.warning(f"[xpoz] \"{kw}\" 에러: {err_msg}")
                yield _make_progress(kw_idx, kw, kw_new, kw_dup, kw_page,
                    f"\"{kw}\" 에러: {err_msg}", error=err_msg)
                continue

        # 모든 키워드 완료
        yield _make_progress(len(keywords)-1, keywords[-1] if keywords else '', 0, 0, 0,
            f"수집 완료! 신규 {new_count}명 / 중복 {dup_count}명 / {total_pages}페이지 / {total_results}건 / 크레딧 {round(credits_used, 1)}",
            done=True)
        try:
            client.close()
        except Exception:
            pass

    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ═══════════════════════════════════════════════════════
# 계정 수집 (자체 — instagrapi + 프록시 + 계정 로테이션)
# 다중 작업 동시 수집 + 계정 풀 공유 + 작업 히스토리
# ═══════════════════════════════════════════════════════

import threading as _threading

_SELF_ACCOUNTS_FILE = os.path.join(os.path.dirname(__file__), "self_accounts.json")
_SELF_JOBS_FILE = os.path.join(os.path.dirname(__file__), "self_jobs.json")
_SELF_PROXIES_FILE = os.path.join(os.path.dirname(__file__), "self_proxies.json")
_SELF_SETTINGS_FILE = os.path.join(os.path.dirname(__file__), "self_settings.json")
_SELF_POSTS_DIR = os.path.join(os.path.dirname(__file__), "self_posts")
os.makedirs(_SELF_POSTS_DIR, exist_ok=True)
_SELF_SESSIONS_DIR = os.path.join(os.path.dirname(__file__), "self_sessions")
os.makedirs(_SELF_SESSIONS_DIR, exist_ok=True)

def _session_path(username: str) -> str:
    return os.path.join(_SELF_SESSIONS_DIR, f"{username}.json")

def _make_challenge_handler(username_hint: str):
    """Instagram 보안 챌린지(SMS/이메일 인증) 발생 시 핸들러.
    SMS/이메일 코드는 자동 수신 불가 → False 반환하여 즉시 실패 처리.
    (TOTP 2FA와는 별개 시스템 — TOTP는 login()의 verification_code로 처리)"""
    def handler(username, choice=None):
        log.warning(f"[challenge] @{username} SMS/이메일 인증 요구 (choice={choice}) — 자동 해제 불가, 수동 필요")
        return False
    return handler

def _setup_client(cl, proxy: str = ""):
    """클라이언트 기본 설정: 지역, 타임존, 프록시, 챌린지 핸들러.
    모든 Client() 생성 후 login 전에 호출해야 함."""
    # 한국 지역 설정 (프록시 위치와 일치해야 차단 방지)
    cl.set_locale("ko_KR")
    cl.set_timezone_offset(9 * 3600)   # UTC+9 (KST)
    cl.set_country("KR")
    cl.set_country_code(82)            # +82
    if proxy:
        cl.set_proxy(proxy)

def _login_with_session(cl, username: str, password: str, totp_secret: str, proxy: str = "") -> bool:
    """세션 파일이 있으면 복원, 없으면 새 로그인. 성공 시 세션 저장.
    핵심 원칙:
    1. UUID/device_settings는 절대 삭제하지 않음 (같은 기기로 인식 유지)
    2. 세션 만료 시 UUID 보존한 채 재로그인
    3. 챌린지 발생 시에도 UUID 보존 (세션 파일 삭제 금지)
    4. locale/timezone/country를 프록시 위치와 일치"""
    import pyotp
    from instagrapi.exceptions import LoginRequired
    spath = _session_path(username)

    # 기본 설정 (지역, 프록시)
    _setup_client(cl, proxy)

    # SMS/이메일 챌린지 핸들러 (자동 불가 → 빠른 실패)
    cl.challenge_code_handler = _make_challenge_handler(username)

    # 1) 저장된 세션 복원 시도
    if os.path.exists(spath):
        try:
            session = cl.load_settings(spath)
            if session:
                cl.set_settings(session)
                # set_settings 후 프록시/지역 재설정 (세션 값으로 덮어쓰이므로)
                _setup_client(cl, proxy)
            # 세션 복원 로그인 — 세션이 살아있으면 TOTP 없이 통과,
            # 만료 시 TwoFactorRequired 방지를 위해 TOTP도 함께 전달
            totp_first = pyotp.TOTP(totp_secret).now() if totp_secret else ""
            cl.login(username, password, verification_code=totp_first)
            # 세션 유효성 검증
            try:
                cl.account_info()
            except LoginRequired:
                log.info(f"[session] @{username} 세션 만료 — UUID 유지 재로그인")
                old = cl.get_settings()
                cl.set_settings({})
                cl.set_uuids(old["uuids"])  # 디바이스 UUID 유지
                _setup_client(cl, proxy)
                totp = pyotp.TOTP(totp_secret).now() if totp_secret else ""
                cl.login(username, password, verification_code=totp)
            cl.dump_settings(spath)
            log.info(f"[session] @{username} 세션 복원 성공")
            return True
        except Exception as e:
            err_str = str(e).lower()
            log.info(f"[session] @{username} 세션 복원 실패: {str(e)[:80]} — UUID 보존 재로그인")
            # ★ 챌린지로 실패해도 세션 파일 삭제 금지!
            # UUID를 보존한 채 재로그인 시도 (새 기기로 인식되면 더 심한 챌린지 유발)
            try:
                old = cl.get_settings()
                saved_uuids = old.get("uuids", {})
                if saved_uuids:
                    cl.set_settings({})
                    cl.set_uuids(saved_uuids)
                    _setup_client(cl, proxy)
                    log.info(f"[session] @{username} UUID 보존 — 같은 기기로 재로그인")
            except Exception:
                _setup_client(cl, proxy)

    # 2) 새 로그인 (TOTP verification_code로 2FA 처리)
    totp = pyotp.TOTP(totp_secret).now() if totp_secret else ""
    cl.login(username, password, verification_code=totp)
    cl.dump_settings(spath)
    log.info(f"[session] @{username} 새 로그인 + 세션 저장")
    return True

def _save_self_posts(job_id: str, hashtag: str, posts: list):
    """게시물을 해시태그별 JSON 파일에 추가 저장"""
    import hashlib
    safe_tag = hashtag.replace("/", "_").replace("\\", "_")
    fpath = os.path.join(_SELF_POSTS_DIR, f"{safe_tag}.json")
    existing = []
    if os.path.exists(fpath):
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass
    existing_ids = {p.get("media_id") for p in existing}
    added = 0
    for p in posts:
        if p.get("media_id") and p["media_id"] not in existing_ids:
            existing.append(p)
            existing_ids.add(p["media_id"])
            added += 1
    if added:
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False)
    return added

def _load_self_posts_tags() -> list:
    """저장된 해시태그 목록 반환"""
    tags = []
    if os.path.exists(_SELF_POSTS_DIR):
        for fname in sorted(os.listdir(_SELF_POSTS_DIR)):
            if fname.endswith(".json"):
                tag = fname[:-5]
                fpath = os.path.join(_SELF_POSTS_DIR, fname)
                try:
                    size = os.path.getsize(fpath)
                    with open(fpath, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    tags.append({"tag": tag, "count": len(data), "size_mb": round(size / 1048576, 1)})
                except Exception:
                    tags.append({"tag": tag, "count": 0, "size_mb": 0})
    return tags

def _load_self_posts_by_tag(tag: str, offset: int = 0, limit: int = 60) -> tuple:
    """특정 해시태그 게시물 로드 (페이징)"""
    safe_tag = tag.replace("/", "_").replace("\\", "_")
    fpath = os.path.join(_SELF_POSTS_DIR, f"{safe_tag}.json")
    if not os.path.exists(fpath):
        return [], 0
    try:
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
        total = len(data)
        # 최신순 정렬
        data.sort(key=lambda x: x.get("taken_at", 0), reverse=True)
        return data[offset:offset+limit], total
    except Exception:
        return [], 0

_DEFAULT_SETTINGS = {
    "page_delay_min": 25,
    "page_delay_max": 35,
    "rest_interval": 10,
    "rest_delay_min": 60,
    "rest_delay_max": 90,
    "max_pages_per_account": 20,
    "account_switch_delay": 10,
    "login_delay_min": 2,
    "login_delay_max": 5,
}

# ─── 블랙리스트 (수집 필터) ─────────────────────────────────────
_SELF_BLACKLIST_FILE = os.path.join(os.path.dirname(__file__), "self_blacklist.json")

_DEFAULT_BLACKLIST = {
    # full_name(프로필 이름)에 포함되면 수집 제외
    "name_keywords": [
        "이슈", "뉴스", "유머", "예능", "맞팔", "선팔",
        "소식", "핫토픽", "아이돌", "드라마", "연예",
        "사건사고", "블랙박스", "매거진", "엔터",
    ],
    # username에 포함되면 수집 제외
    "username_keywords": [
        "news", "issue", "celeb", "enter.",
    ],
    # 활성화 여부
    "enabled": True,
}

def _load_blacklist() -> dict:
    if os.path.exists(_SELF_BLACKLIST_FILE):
        try:
            with open(_SELF_BLACKLIST_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return {**_DEFAULT_BLACKLIST, **data}
        except Exception:
            pass
    _save_blacklist(_DEFAULT_BLACKLIST)
    return dict(_DEFAULT_BLACKLIST)

def _save_blacklist(bl: dict):
    with open(_SELF_BLACKLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(bl, f, ensure_ascii=False, indent=2)

def _is_blacklisted(username: str, full_name: str, blacklist: dict) -> str:
    """블랙리스트 키워드 매칭. 매칭되면 사유 반환, 아니면 빈 문자열."""
    if not blacklist.get("enabled", True):
        return ""
    fn_lower = full_name.lower()
    un_lower = username.lower()
    for kw in blacklist.get("name_keywords", []):
        if kw.lower() in fn_lower:
            return f"이름:{kw}"
    for kw in blacklist.get("username_keywords", []):
        if kw.lower() in un_lower:
            return f"아이디:{kw}"
    return ""

def _load_self_settings() -> dict:
    if os.path.exists(_SELF_SETTINGS_FILE):
        try:
            with open(_SELF_SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                merged = {**_DEFAULT_SETTINGS, **data}
                return merged
        except Exception:
            pass
    _save_self_settings(_DEFAULT_SETTINGS)
    return dict(_DEFAULT_SETTINGS)

def _save_self_settings(settings: dict):
    with open(_SELF_SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

def _load_self_accounts() -> list:
    if os.path.exists(_SELF_ACCOUNTS_FILE):
        try:
            with open(_SELF_ACCOUNTS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
        except Exception:
            pass
    return []

def _save_self_accounts(accounts: list):
    with open(_SELF_ACCOUNTS_FILE, "w", encoding="utf-8") as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)

def _load_self_jobs() -> list:
    if os.path.exists(_SELF_JOBS_FILE):
        try:
            with open(_SELF_JOBS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
        except Exception:
            pass
    return []

def _save_self_jobs(jobs: list):
    with open(_SELF_JOBS_FILE, "w", encoding="utf-8") as f:
        json.dump(jobs, f, ensure_ascii=False, indent=2)

def _load_self_proxies() -> list:
    if os.path.exists(_SELF_PROXIES_FILE):
        try:
            with open(_SELF_PROXIES_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
        except Exception:
            pass
    # 초기화: .env 프록시가 있으면 자동 등록
    env_proxy = os.getenv("IPROYAL_PROXY", "")
    if env_proxy:
        default = [{"name": "IPRoyal", "url": env_proxy, "active": True}]
        _save_self_proxies(default)
        return default
    return []

def _save_self_proxies(proxies: list):
    with open(_SELF_PROXIES_FILE, "w", encoding="utf-8") as f:
        json.dump(proxies, f, ensure_ascii=False, indent=2)

def _get_active_proxy() -> str:
    """활성 프록시 URL 반환"""
    proxies = _load_self_proxies()
    for p in proxies:
        if p.get("active"):
            return p.get("url", "")
    return ""

# 다중 작업 상태 관리
_self_jobs_running = {}       # {job_id: {progress dict, stop_requested, thread}}
_self_account_claims = {}     # {username: job_id} — 현재 사용 중인 계정
_self_lock = _threading.Lock()
_self_last_account_idx = 0    # 마지막 사용 계정 인덱스 (전역, 모든 작업 공유)

def _get_last_account_idx() -> int:
    """settings에서 마지막 사용 계정 인덱스 로드"""
    global _self_last_account_idx
    s = _load_self_settings()
    _self_last_account_idx = s.get("last_account_idx", 0)
    return _self_last_account_idx

def _save_last_account_idx(idx: int):
    """마지막 사용 계정 인덱스 저장"""
    global _self_last_account_idx
    _self_last_account_idx = idx
    s = _load_self_settings()
    s["last_account_idx"] = idx
    _save_self_settings(s)

# 서버 시작 시 고아 running 작업 정리
def _cleanup_orphan_jobs():
    jobs = _load_self_jobs()
    changed = False
    for j in jobs:
        if j.get("status") == "running":
            j["status"] = "stopped"
            j["end_time"] = datetime.now(timezone.utc).isoformat()
            j["error"] = "서버 재시작으로 중지됨"
            changed = True
    if changed:
        _save_self_jobs(jobs)
_cleanup_orphan_jobs()

def _claim_account(job_id: str, accounts: list, start_from: int = 0) -> int:
    """사용 가능한 계정 인덱스 반환. 없으면 -1."""
    with _self_lock:
        for i in range(len(accounts)):
            idx = (start_from + i) % len(accounts)
            uname = accounts[idx]["username"]
            status = accounts[idx].get("status", "idle")
            if uname not in _self_account_claims and status not in ("blocked", "login_failed"):
                _self_account_claims[uname] = job_id
                return idx
    return -1

def _release_account(username: str):
    with _self_lock:
        _self_account_claims.pop(username, None)

def _get_running_jobs_summary():
    """실행 중인 작업 요약 (SSE 없이 상태 확인용)"""
    result = []
    with _self_lock:
        for jid, jstate in _self_jobs_running.items():
            p = jstate.get("progress", {})
            result.append({
                "job_id": jid,
                "hashtag": p.get("hashtag", ""),
                "new": p.get("new", 0),
                "dup": p.get("dup", 0),
                "posts": p.get("posts", 0),
                "pages": p.get("pages", 0),
                "status": p.get("status", ""),
                "current_account": p.get("current_account", ""),
            })
    return result

@app.get("/self-collect", response_class=HTMLResponse)
def self_collect_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return RedirectResponse("/login", 302)
    try:
        accounts = _load_self_accounts()
        proxies = _load_self_proxies()
        jobs_history = _load_self_jobs()[-50:]
        running_jobs = _get_running_jobs_summary()
        # 계정 상태 집계
        acc_stats = {"total": len(accounts), "idle": 0, "active": 0, "blocked": 0, "failed": 0, "resting": 0}
        for a in accounts:
            st = a.get("status", "idle")
            if st in ("idle", ""): acc_stats["idle"] += 1
            elif st in ("active", "logging_in"): acc_stats["active"] += 1
            elif st == "blocked": acc_stats["blocked"] += 1
            elif st == "login_failed": acc_stats["failed"] += 1
            elif st == "resting": acc_stats["resting"] += 1
            else: acc_stats["idle"] += 1
        claimed = dict(_self_account_claims)  # {username: job_id}
        settings = _load_self_settings()
        blacklist = _load_blacklist()
        return templates.TemplateResponse("self_collect.html", {
            "request": request, "user": user, "active": "self_collect",
            "accounts": accounts, "proxies": proxies,
            "jobs_history": jobs_history,
            "running_jobs": running_jobs,
            "acc_stats": acc_stats,
            "claimed_accounts": claimed,
            "settings": settings,
            "blacklist": blacklist,
        })
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        log.error(f"self-collect 페이지 에러: {err}")
        return HTMLResponse(f"<pre style='color:red;padding:20px'>{err}</pre>", status_code=500)

@app.get("/api/self-collect/accounts")
def self_collect_get_accounts(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    accounts = _load_self_accounts()
    safe = []
    for a in accounts:
        uname = a["username"]
        claimed_by = _self_account_claims.get(uname)
        safe.append({
            "username": uname,
            "status": a.get("status", "idle"),
            "error": a.get("error", ""),
            "claimed_by": claimed_by,
        })
    return JSONResponse({"accounts": safe})

@app.post("/api/self-collect/accounts/import")
async def self_collect_import_accounts(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    body = await request.json()
    text = body.get("text", "")
    accounts = _load_self_accounts()
    existing = {a["username"] for a in accounts}
    added = 0
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(":")
        if len(parts) >= 3:
            uname = parts[0].strip()
            if uname not in existing:
                accounts.append({
                    "username": uname,
                    "password": parts[1].strip(),
                    "totp_secret": parts[2].strip(),
                    "status": "idle",
                    "error": "",
                })
                existing.add(uname)
                added += 1
    _save_self_accounts(accounts)
    return JSONResponse({"message": f"{added}개 계정 추가 (총 {len(accounts)}개)", "total": len(accounts)})

@app.delete("/api/self-collect/accounts/{username}")
def self_collect_delete_account(username: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    if username in _self_account_claims:
        return JSONResponse({"error": f"@{username}은 현재 사용 중 (작업: {_self_account_claims[username]})"}, 400)
    accounts = _load_self_accounts()
    accounts = [a for a in accounts if a["username"] != username]
    _save_self_accounts(accounts)
    return JSONResponse({"message": f"@{username} 삭제", "total": len(accounts)})

def _translate_insta_err(raw: str) -> str:
    el = raw.lower()
    if "recaptcha" in el or "captcha" in el:
        return "캡차 인증 필요 — 수동 해제 필요"
    if "selfie" in el:
        return "셀피 인증 필요 — 수동 해제 필요"
    if "challengeresolve" in el or "challenge" in el:
        return "보안 인증 필요 — 자동 해제 실패 시 수동 확인"
    if "can't find an account" in el or "find an account" in el or "invalid_user" in el or "user_not_found" in el:
        return "삭제/정지된 계정 — Instagram에 존재하지 않음"
    if "we can send you an email" in el or "get back into your account" in el:
        return "계정 복구 필요 — Instagram 이메일 인증 필요"
    if "login_required" in el:
        return "로그인 만료 — 재로그인 필요"
    if "feedback_required" in el:
        return "계정 활동 제한 — Instagram이 자동화 감지"
    if "consent_required" in el:
        return "약관 동의 필요 — Instagram 앱에서 직접 로그인 필요"
    if "checkpoint" in el:
        return "보안 체크포인트 — Instagram 앱에서 인증 필요"
    if "bad_password" in el or "incorrect" in el:
        return "비밀번호 오류"
    if "two_factor" in el or "2fa" in el or "verification" in el:
        return "2FA 인증 실패 — TOTP 코드 확인"
    if "rate_limit" in el or "please wait" in el:
        return "요청 제한 — 잠시 후 재시도"
    if "connection" in el or "timeout" in el:
        return "연결 오류 — 프록시/네트워크 확인"
    return raw[:80]

# 기존 계정 에러 메시지 한글화 (서버 시작 시)
def _retranslate_account_errors():
    accounts = _load_self_accounts()
    changed = False
    for a in accounts:
        err = a.get("error", "")
        if not err:
            continue
        # 코드 에러 메시지 정리
        if "name '" in err and "not defined" in err:
            a["error"] = "시스템 오류 — 재확인 필요"
            a["status"] = "idle"
            changed = True
            continue
        # 영어 에러 한글화
        if any(c.isascii() and c.isalpha() for c in err[:20]):
            translated = _translate_insta_err(err)
            if translated != err:
                a["error"] = translated
                changed = True
    if changed:
        _save_self_accounts(accounts)
_retranslate_account_errors()

@app.post("/api/self-collect/accounts/{username}/check")
def self_collect_check_account(username: str, use_proxy: int = Query(default=1), session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    accounts = _load_self_accounts()
    acc = None
    for a in accounts:
        if a["username"] == username:
            acc = a
            break
    if not acc:
        return JSONResponse({"error": "계정을 찾을 수 없음"}, 404)

    from instagrapi import Client
    proxy = _get_active_proxy() if use_proxy else ""
    try:
        cl = Client()
        cl.delay_range = [2, 5]
        _login_with_session(cl, username, acc["password"], acc.get("totp_secret", ""), proxy)
        # 성공 — 세션 저장 + 상태 업데이트
        acc["status"] = "idle"
        acc["error"] = ""
        _save_self_accounts(accounts)
        try:
            cl.logout()
        except Exception:
            pass
        return JSONResponse({"ok": True, "message": "로그인 성공"})
    except Exception as e:
        err_raw = str(e)[:200]
        # 에러 번역
        el = err_raw.lower()
        err_kr = _translate_insta_err(err_raw)

        acc["status"] = "login_failed"
        acc["error"] = err_kr
        _save_self_accounts(accounts)
        return JSONResponse({"ok": False, "error": err_kr})

@app.post("/api/self-collect/accounts/reactivate-all")
def self_collect_reactivate_all(session_id: Optional[str] = Cookie(default=None)):
    """차단/실패 계정 전체 재활성화"""
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    accounts = _load_self_accounts()
    count = 0
    for a in accounts:
        if a.get("status") in ("blocked", "login_failed"):
            a["status"] = "idle"
            a["error"] = ""
            a.pop("blocked_at", None)
            count += 1
            # 세션 파일 삭제 대신 쿠키만 초기화 (UUID/디바이스 지문 보존)
            spath = _session_path(a["username"])
            if os.path.exists(spath):
                try:
                    import json as _json
                    with open(spath, "r") as f:
                        sess = _json.load(f)
                    saved_uuids = sess.get("uuids", {})
                    cleaned = {"uuids": saved_uuids} if saved_uuids else {}
                    with open(spath, "w") as f:
                        _json.dump(cleaned, f)
                except Exception:
                    pass
    if count:
        _save_self_accounts(accounts)
    log.info(f"[self-collect] 전체 재활성화: {count}개 계정")
    return JSONResponse({"ok": True, "count": count, "message": f"{count}개 계정 재활성화"})

@app.post("/api/self-collect/accounts/{username}/reactivate")
def self_collect_reactivate_account(username: str, session_id: Optional[str] = Cookie(default=None)):
    """차단/실패 계정을 재활성화 (상태 초기화)"""
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    accounts = _load_self_accounts()
    for a in accounts:
        if a["username"] == username:
            a["status"] = "idle"
            a["error"] = ""
            a.pop("blocked_at", None)
            _save_self_accounts(accounts)
            # 세션 파일에서 쿠키만 초기화 (UUID/디바이스 지문 보존)
            spath = _session_path(username)
            if os.path.exists(spath):
                try:
                    import json as _json
                    with open(spath, "r") as f:
                        sess = _json.load(f)
                    saved_uuids = sess.get("uuids", {})
                    cleaned = {"uuids": saved_uuids} if saved_uuids else {}
                    with open(spath, "w") as f:
                        _json.dump(cleaned, f)
                except Exception:
                    pass
            log.info(f"[self-collect] @{username} 재활성화")
            return JSONResponse({"ok": True, "message": f"@{username} 재활성화 완료"})
    return JSONResponse({"error": "계정을 찾을 수 없음"}, 404)

@app.post("/api/self-collect/accounts/{username}/unblock")
def self_collect_unblock_account(username: str, session_id: Optional[str] = Cookie(default=None)):
    """크롬 시크릿 창에서 Instagram 로그인 → 사용자가 수동으로 인증 해제"""
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    accounts = _load_self_accounts()
    acc = None
    for a in accounts:
        if a["username"] == username:
            acc = a
            break
    if not acc:
        return JSONResponse({"error": "계정을 찾을 수 없음"}, 404)

    import threading as _thr

    def _open_chrome():
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from webdriver_manager.chrome import ChromeDriverManager

            opts = Options()
            opts.add_argument("--incognito")
            opts.add_argument("--start-maximized")
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)
            # 프록시 설정 — 해제도 프록시 IP로 해야 재챌린지 방지
            _proxy_url = _get_active_proxy()
            _proxy_ext_path = None
            if _proxy_url:
                from urllib.parse import urlparse
                import zipfile, tempfile
                pp = urlparse(_proxy_url)
                if pp.username and pp.password:
                    # 인증 프록시 → Chrome MV3 확장으로 자동 인증
                    manifest = json.dumps({"name":"Proxy Auth","version":"1.0.0","manifest_version":3,"permissions":["proxy","webRequest","webRequestAuthProvider"],"host_permissions":["<all_urls>"],"background":{"service_worker":"bg.js"}})
                    bg_js = f'''chrome.proxy.settings.set({{value:{{mode:"fixed_servers",rules:{{singleProxy:{{scheme:"{pp.scheme or "http"}",host:"{pp.hostname}",port:{pp.port}}},bypassList:["localhost"]}}}},scope:"regular"}});
chrome.webRequest.onAuthRequired.addListener(function(d){{return{{authCredentials:{{username:"{pp.username}",password:"{pp.password}"}}}}}},{{urls:["<all_urls>"]}},["blocking"]);'''
                    _proxy_ext_path = os.path.join(tempfile.gettempdir(), f"proxy_auth_{username}.zip")
                    with zipfile.ZipFile(_proxy_ext_path, 'w') as zf:
                        zf.writestr("manifest.json", manifest)
                        zf.writestr("bg.js", bg_js)
                    opts.add_extension(_proxy_ext_path)
                    # 인증 프록시 확장 사용 시 incognito 불가 → 제거
                    opts._arguments = [a for a in opts._arguments if a != "--incognito"]
                    log.info(f"[unblock] 크롬 인증 프록시 확장 설정: {pp.hostname}:{pp.port}")
                else:
                    opts.add_argument(f"--proxy-server={pp.scheme}://{pp.hostname}:{pp.port}")
                    log.info(f"[unblock] 크롬 프록시 설정: {pp.hostname}:{pp.port}")

            svc = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=svc, options=opts)
            import time as _t

            driver.get("https://www.instagram.com/accounts/login/")
            _t.sleep(3)

            # 쿠키 동의 팝업 닫기 (있으면)
            try:
                cookie_btn = driver.find_element(By.XPATH, "//button[contains(text(),'Allow')]|//button[contains(text(),'Accept')]|//button[contains(text(),'허용')]|//button[contains(text(),'모두 허용')]")
                cookie_btn.click()
                _t.sleep(1)
            except Exception:
                pass

            # 로그인 폼 대기
            wait = WebDriverWait(driver, 20)
            uname_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='username']")))
            _t.sleep(1)

            # 클릭 후 send_keys (한 글자씩)
            from selenium.webdriver.common.keys import Keys
            uname_input.click()
            _t.sleep(0.3)
            uname_input.send_keys(Keys.CONTROL + "a")
            uname_input.send_keys(Keys.DELETE)
            for ch in acc["username"]:
                uname_input.send_keys(ch)
                _t.sleep(0.05)
            _t.sleep(0.5)

            pwd_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='password']")))
            pwd_input.click()
            _t.sleep(0.3)
            pwd_input.send_keys(Keys.CONTROL + "a")
            pwd_input.send_keys(Keys.DELETE)
            for ch in acc["password"]:
                pwd_input.send_keys(ch)
                _t.sleep(0.05)
            _t.sleep(1)

            # 로그인 버튼 클릭
            login_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']")))
            login_btn.click()
            log.info(f"[unblock] @{username} 크롬 시크릿 창 열림 — 사용자 수동 인증 대기")
            # 브라우저는 열어둔 채로 유지 (사용자가 직접 닫음)
        except Exception as e:
            log.error(f"[unblock] @{username} 크롬 열기 실패: {e}")

    _thr.Thread(target=_open_chrome, daemon=True).start()
    # 2FA 코드 생성
    totp_code = ""
    totp_secret = acc.get("totp_secret", "")
    if totp_secret:
        import pyotp
        totp_code = pyotp.TOTP(totp_secret).now()
    return JSONResponse({"ok": True, "message": f"@{username} 크롬 시크릿 창 열는 중...", "totp_code": totp_code})

@app.get("/api/self-collect/accounts/{username}/totp")
def self_collect_account_totp(username: str, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    accounts = _load_self_accounts()
    for a in accounts:
        if a["username"] == username:
            secret = a.get("totp_secret", "")
            if secret:
                import pyotp
                return JSONResponse({"code": pyotp.TOTP(secret).now()})
            return JSONResponse({"code": ""})
    return JSONResponse({"error": "계정 없음"}, 404)

# ── 프록시 관리 API ──
@app.get("/api/self-collect/proxies")
def self_collect_get_proxies(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    return JSONResponse({"proxies": _load_self_proxies()})

@app.post("/api/self-collect/proxies")
async def self_collect_add_proxy(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    body = await request.json()
    name = body.get("name", "").strip()
    url = body.get("url", "").strip()
    if not name or not url:
        return JSONResponse({"error": "이름과 URL 필수"}, 400)
    proxies = _load_self_proxies()
    proxies.append({"name": name, "url": url, "active": len(proxies) == 0})
    _save_self_proxies(proxies)
    return JSONResponse({"message": f"{name} 추가", "proxies": proxies})

@app.put("/api/self-collect/proxies/{idx}")
async def self_collect_update_proxy(idx: int, request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    body = await request.json()
    proxies = _load_self_proxies()
    if idx < 0 or idx >= len(proxies):
        return JSONResponse({"error": "잘못된 인덱스"}, 400)
    if "name" in body:
        proxies[idx]["name"] = body["name"].strip()
    if "url" in body:
        proxies[idx]["url"] = body["url"].strip()
    if "active" in body:
        # 하나만 활성
        if body["active"]:
            for i, p in enumerate(proxies):
                p["active"] = (i == idx)
        else:
            proxies[idx]["active"] = False
    _save_self_proxies(proxies)
    return JSONResponse({"message": "수정 완료", "proxies": proxies})

@app.delete("/api/self-collect/proxies/{idx}")
def self_collect_delete_proxy(idx: int, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    proxies = _load_self_proxies()
    if idx < 0 or idx >= len(proxies):
        return JSONResponse({"error": "잘못된 인덱스"}, 400)
    removed = proxies.pop(idx)
    # 삭제된 게 활성이었으면 첫 번째를 활성으로
    if removed.get("active") and proxies:
        proxies[0]["active"] = True
    _save_self_proxies(proxies)
    return JSONResponse({"message": f"{removed['name']} 삭제", "proxies": proxies})

@app.post("/api/self-collect/start")
async def self_collect_start(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)

    body = await request.json()
    hashtag = body.get("hashtag", "").strip().lstrip("#")
    search_type = body.get("search_type", "recent")
    target_users = int(body.get("target_users", 1000))
    location = body.get("location", "").strip()

    if not hashtag and not location:
        return JSONResponse({"error": "해시태그 또는 위치 필수"}, 400)

    accounts = _load_self_accounts()
    if not accounts:
        return JSONResponse({"error": "등록된 계정 없음"}, 400)

    # 사용 가능한 계정 확인
    available = sum(1 for a in accounts if a["username"] not in _self_account_claims and a.get("status") not in ("blocked", "login_failed"))
    if available == 0:
        return JSONResponse({"error": "사용 가능한 계정 없음 (모두 사용 중이거나 차단)"}, 400)

    proxy = _get_active_proxy()
    job_id = f"j{int(time.time()*1000)%1000000:06d}"

    # 작업 히스토리에 기록
    job_record = {
        "id": job_id,
        "hashtag": hashtag,
        "search_type": search_type,
        "location": location,
        "target_users": target_users,
        "start_time": datetime.now(timezone.utc).isoformat(),
        "end_time": None,
        "status": "running",
        "new_count": 0,
        "dup_count": 0,
        "total_posts": 0,
        "total_pages": 0,
        "error": None,
        "users": [],
    }
    jobs = _load_self_jobs()
    jobs.append(job_record)
    _save_self_jobs(jobs)

    # 실행 상태 등록
    job_state = {
        "stop_requested": False,
        "progress": {
            "job_id": job_id, "hashtag": hashtag,
            "new": 0, "dup": 0, "posts": 0, "pages": 0,
            "status": "시작 중...", "current_account": "",
            "account_index": 0, "total_accounts": len(accounts),
        },
    }
    with _self_lock:
        _self_jobs_running[job_id] = job_state

    t = _threading.Thread(
        target=_self_collect_worker,
        args=(job_id, hashtag, search_type, target_users, proxy, location),
        daemon=True,
    )
    job_state["thread"] = t
    t.start()
    return JSONResponse({"message": "수집 시작", "job_id": job_id, "hashtag": hashtag})

@app.post("/api/self-collect/stop")
async def self_collect_stop(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    body = await request.json()
    job_id = body.get("job_id", "")
    if not job_id:
        # 전체 중지
        with _self_lock:
            for jid in _self_jobs_running:
                _self_jobs_running[jid]["stop_requested"] = True
        return JSONResponse({"message": "전체 중지 요청"})
    if job_id in _self_jobs_running:
        _self_jobs_running[job_id]["stop_requested"] = True
        return JSONResponse({"message": f"{job_id} 중지 요청"})
    return JSONResponse({"error": "작업을 찾을 수 없음"}, 404)

@app.get("/api/self-collect/stream")
def self_collect_stream(job_id: str = "", session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)

    def stream():
        import copy
        while True:
            with _self_lock:
                if job_id and job_id in _self_jobs_running:
                    progress = copy.deepcopy(_self_jobs_running[job_id]["progress"])
                    # users는 한 번 보내면 클리어 (중복 전송 방지)
                    _self_jobs_running[job_id]["progress"].pop("users", None)
                elif not job_id:
                    # 전체 작업 요약
                    all_progress = []
                    for jid, js in _self_jobs_running.items():
                        all_progress.append(copy.deepcopy(js["progress"]))
                    progress = {"jobs": all_progress, "count": len(all_progress)}
                else:
                    yield f"data: {json.dumps({'done': True, 'status': '작업 종료'}, ensure_ascii=False)}\n\n"
                    return
            yield f"data: {json.dumps(progress, ensure_ascii=False)}\n\n"
            if isinstance(progress, dict) and progress.get("done"):
                break
            # 모든 작업이 끝났는지 확인
            if not job_id:
                with _self_lock:
                    if not _self_jobs_running:
                        yield f"data: {json.dumps({'jobs': [], 'count': 0, 'all_done': True}, ensure_ascii=False)}\n\n"
                        break
            time.sleep(2)

    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

@app.get("/api/self-collect/status")
def self_collect_status(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    with _self_lock:
        running = []
        for jid, js in _self_jobs_running.items():
            running.append({"job_id": jid, "progress": js["progress"]})
        return JSONResponse({
            "running_count": len(running),
            "jobs": running,
            "claimed_accounts": dict(_self_account_claims),
        })

@app.get("/api/self-collect/jobs")
def self_collect_jobs_list(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    jobs = _load_self_jobs()
    jobs.reverse()  # 최신순
    return JSONResponse({"jobs": jobs[:100]})

@app.get("/api/self-collect/settings")
def self_collect_get_settings(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    return JSONResponse(_load_self_settings())

@app.post("/api/self-collect/settings")
async def self_collect_save_settings(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    body = await request.json()
    settings = _load_self_settings()
    for k in _DEFAULT_SETTINGS:
        if k in body:
            try:
                settings[k] = int(body[k])
            except (ValueError, TypeError):
                pass
    _save_self_settings(settings)
    return JSONResponse({"ok": True, "settings": settings})

# ─── 블랙리스트 API ──────────────────────────────────────────────
@app.get("/api/self-collect/blacklist")
def self_collect_get_blacklist(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    return JSONResponse(_load_blacklist())

@app.post("/api/self-collect/blacklist")
async def self_collect_save_blacklist(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    body = await request.json()
    bl = _load_blacklist()
    if "name_keywords" in body and isinstance(body["name_keywords"], list):
        bl["name_keywords"] = [kw.strip() for kw in body["name_keywords"] if kw.strip()]
    if "username_keywords" in body and isinstance(body["username_keywords"], list):
        bl["username_keywords"] = [kw.strip() for kw in body["username_keywords"] if kw.strip()]
    if "enabled" in body:
        bl["enabled"] = bool(body["enabled"])
    _save_blacklist(bl)
    return JSONResponse({"ok": True, "blacklist": bl})

# ─── 게시물 탐색 페이지 ──────────────────────────────────────────
@app.get("/self-collect/posts", response_class=HTMLResponse)
def self_collect_posts_page(request: Request, tag: str = "", page: int = 1, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return RedirectResponse("/login", 302)
    tags_list = _load_self_posts_tags()
    posts = []
    total = 0
    per_page = 60
    if tag:
        posts, total = _load_self_posts_by_tag(tag, offset=(page - 1) * per_page, limit=per_page)
    total_pages = (total + per_page - 1) // per_page if total else 0
    return templates.TemplateResponse("self_posts.html", {
        "request": request, "user": user, "active": "self_posts",
        "tags_list": tags_list, "current_tag": tag,
        "posts": posts, "total": total,
        "page": page, "total_pages": total_pages,
    })

@app.get("/api/self-collect/posts")
def self_collect_posts_api(tag: str = "", page: int = 1, limit: int = 60, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user:
        return JSONResponse({"error": "인증 필요"}, 403)
    if not tag:
        return JSONResponse({"posts": [], "total": 0, "tags": _load_self_posts_tags()})
    posts, total = _load_self_posts_by_tag(tag, offset=(page - 1) * limit, limit=limit)
    return JSONResponse({"posts": posts, "total": total, "page": page})

def _self_collect_worker(job_id, hashtag, search_type, target_users, proxy, location=""):
    """백그라운드 크롤링 워커 — 계정 풀 공유 + 선제 로테이션 + 작업 히스토리"""
    import pyotp, random
    from instagrapi import Client
    from database import upsert_influencer, get_existing_pks, get_banned_pks, ban_influencer

    job_state = _self_jobs_running.get(job_id)
    if not job_state:
        return
    progress = job_state["progress"]

    def _update(**kwargs):
        with _self_lock:
            progress.update(kwargs)
        # 신규/중복/게시물/페이지 값이 있으면 히스토리 파일에도 실시간 저장
        if any(k in kwargs for k in ("new", "dup", "posts", "pages")):
            try:
                jobs = _load_self_jobs()
                for j in jobs:
                    if j["id"] == job_id:
                        if "new" in kwargs: j["new_count"] = kwargs["new"]
                        if "dup" in kwargs: j["dup_count"] = kwargs["dup"]
                        if "posts" in kwargs: j["total_posts"] = kwargs["posts"]
                        if "pages" in kwargs: j["total_pages"] = kwargs["pages"]
                        break
                _save_self_jobs(jobs)
            except Exception:
                pass

    def _update_account_status(uname, status, error=""):
        accounts = _load_self_accounts()
        for a in accounts:
            if a["username"] == uname:
                a["status"] = status
                a["error"] = error
                if status == "blocked":
                    a["blocked_at"] = datetime.now(timezone.utc).isoformat()
                break
        _save_self_accounts(accounts)

    def _finish_job(new_count, dup_count, total_posts, total_pages, error=None):
        """작업 히스토리 업데이트"""
        jobs = _load_self_jobs()
        for j in jobs:
            if j["id"] == job_id:
                j["end_time"] = datetime.now(timezone.utc).isoformat()
                j["status"] = "error" if error else ("stopped" if job_state["stop_requested"] else "done")
                j["new_count"] = new_count
                j["dup_count"] = dup_count
                j["total_posts"] = total_posts
                j["total_pages"] = total_pages
                j["error"] = error
                break
        _save_self_jobs(jobs)

    # DB 기존 유저
    existing_pks = set()
    try:
        existing_pks = set(get_existing_pks())
    except Exception:
        pass
    banned_pks = set()
    try:
        banned_pks = set(get_banned_pks())
    except Exception:
        pass

    seen_user_pks = set()
    new_count = 0
    dup_count = 0
    skip_count = 0  # 블랙리스트 필터링 카운트
    total_posts = 0
    total_pages = 0
    cursor = None
    account_pages = 0
    last_account_idx = _get_last_account_idx()  # 마지막 사용 계정 인덱스 (파일에서 복원)
    _settings = _load_self_settings()
    _blacklist = _load_blacklist()
    MAX_PAGES_PER_ACCOUNT = _settings["max_pages_per_account"]
    COOLDOWN_SECONDS = 120
    tag = hashtag
    tab = "recent" if search_type == "recent" else "top"
    is_location_search = (search_type == "location")
    location_pk = None
    current_account_username = None

    def _translate_insta_error(err: str) -> str:
        el = err.lower()
        if "recaptcha" in el or "captcha" in el:
            return "캡차 인증 필요 — 수동 해제 필요"
        if "selfie" in el:
            return "셀피 인증 필요 — 수동 해제 필요"
        if "challengeresolve" in el or "challenge" in el:
            return "보안 인증 — 자동 해제 실패 시 수동 확인"
        if "login_required" in el:
            return "로그인 만료 — 재로그인 필요"
        if "feedback_required" in el:
            return "계정 활동 제한 — Instagram이 자동화 감지"
        if "consent_required" in el:
            return "약관 동의 필요 — Instagram 앱에서 직접 로그인 필요"
        if "checkpoint" in el:
            return "보안 체크포인트 — Instagram 앱에서 인증 필요"
        if "bad_password" in el or "incorrect" in el:
            return "비밀번호 오류"
        if "invalid_user" in el or "user_not_found" in el or "can't find an account" in el or "find an account" in el:
            return "삭제/정지된 계정 — Instagram에 존재하지 않음"
        if "rate_limit" in el or "please wait" in el:
            return "요청 제한 — 잠시 후 재시도"
        if "two_factor" in el or "2fa" in el or "verification" in el:
            return "2FA 인증 실패 — TOTP 코드 확인 필요"
        if "connection" in el or "timeout" in el:
            return "연결 오류 — 프록시/네트워크 확인"
        return err[:80]

    def _login_next(start_from=None):
        """계정 풀에서 사용 가능한 계정 할당 후 로그인. start_from=None이면 마지막 사용 인덱스+1부터"""
        nonlocal current_account_username, last_account_idx
        if start_from is None:
            start_from = last_account_idx + 1
        accounts = _load_self_accounts()
        attempts = 0
        idx = start_from
        while attempts < len(accounts):
            claimed_idx = _claim_account(job_id, accounts, idx)
            if claimed_idx < 0:
                # 모든 계정이 사용 중 — 1초 간격 중지 확인하며 대기
                _update(status=f"[{job_id}] 사용 가능한 계정 대기 중...")
                for _ in range(30):
                    if job_state["stop_requested"]:
                        break
                    time.sleep(1)
                accounts = _load_self_accounts()
                attempts += 1
                continue
            acc = accounts[claimed_idx]
            uname = acc["username"]
            _update(status=f"@{uname} 로그인 중...", current_account=uname, account_index=claimed_idx)
            _update_account_status(uname, "logging_in")
            try:
                cl = Client()
                # delay_range: 모든 API 호출 사이에 랜덤 딜레이 (봇 탐지 방지)
                cl.delay_range = [_settings["login_delay_min"], _settings["login_delay_max"]]
                _login_with_session(cl, uname, acc["password"], acc.get("totp_secret", ""), proxy)
                # 로그인 성공 후 크롤링용 딜레이로 변경 (1~3초 — API 호출마다 자동 적용)
                cl.delay_range = [1, 3]
                log.info(f"[self-collect:{job_id}] @{uname} 로그인 성공")
                _update_account_status(uname, "active")
                current_account_username = uname
                last_account_idx = claimed_idx
                _save_last_account_idx(claimed_idx)
                return cl, uname
            except Exception as e:
                err_raw = str(e)[:150]
                el = err_raw.lower()
                is_challenge_err = any(k in el for k in ["challenge", "checkpoint"])
                is_unsolvable = any(k in el for k in ["recaptcha", "captcha", "selfie"])
                if is_challenge_err and not is_unsolvable:
                    # _login_with_session이 세션삭제+TOTP 재로그인 시도했으나 실패
                    log.warning(f"[self-collect:{job_id}] @{uname} TOTP 재로그인 실패: {err_raw}")
                    err = "TOTP 재로그인 실패 — SMS/이메일 인증 또는 수동 확인 필요"
                elif is_unsolvable:
                    log.warning(f"[self-collect:{job_id}] @{uname} 자동 해제 불가: {err_raw}")
                    err = _translate_insta_error(err_raw)
                else:
                    log.warning(f"[self-collect:{job_id}] @{uname} 로그인 실패: {err_raw}")
                    err = _translate_insta_error(err_raw)
                _update_account_status(uname, "login_failed", err)
                _release_account(uname)
                idx = claimed_idx + 1
                attempts += 1
                time.sleep(3)
        return None, None

    def _release_current():
        nonlocal current_account_username
        if current_account_username:
            _release_account(current_account_username)
            current_account_username = None

    # 첫 계정 로그인
    client, acc_uname = _login_next()  # 마지막 사용 계정 다음부터
    if not client:
        _update(done=True, status="사용 가능한 계정 없음", error="no_accounts")
        _finish_job(0, 0, 0, 0, "no_accounts")
        with _self_lock:
            _self_jobs_running.pop(job_id, None)
        return

    # 위치 검색이면 location PK 찾기
    if is_location_search and location:
        _update(status=f"@{acc_uname}로 위치 '{location}' 검색 중...")
        try:
            loc_result = client.private_request("location_search/", params={"search_query": location, "latitude": 0, "longitude": 0})
            venues = loc_result.get("venues", [])
            if venues:
                location_pk = venues[0].get("external_id") or venues[0].get("pk")
                loc_name = venues[0].get("name", location)
                _update(status=f"위치 '{loc_name}' (PK:{location_pk}) 발견 — 수집 시작")
                log.info(f"[self-collect:{job_id}] 위치 '{loc_name}' PK={location_pk}")
            else:
                _update(done=True, status=f"위치 '{location}' 검색 결과 없음", error="위치를 찾을 수 없습니다")
                _finish_job(0, 0, 0, 0, f"위치 '{location}' 검색 결과 없음")
                _release_current()
                with _self_lock:
                    _self_jobs_running.pop(job_id, None)
                return
        except Exception as e:
            _update(done=True, status=f"위치 검색 실패: {str(e)[:80]}", error=str(e)[:100])
            _finish_job(0, 0, 0, 0, f"위치 검색 실패: {str(e)[:100]}")
            _release_current()
            with _self_lock:
                _self_jobs_running.pop(job_id, None)
            return
    else:
        _update(status=f"@{acc_uname}로 수집 시작 — #{tag}")

    try:
        while new_count < target_users and not job_state["stop_requested"]:
            total_pages += 1

            # 안전 딜레이 (설정값 사용) — 1초 간격으로 중지 확인
            if total_pages > 1:
                if total_pages % _settings["rest_interval"] == 0:
                    rest = random.uniform(_settings["rest_delay_min"], _settings["rest_delay_max"])
                    _update(status=f"p{total_pages} 쉬는 중... ({int(rest)}초)")
                    for _ in range(int(rest)):
                        if job_state["stop_requested"]:
                            break
                        time.sleep(1)
                else:
                    delay = random.uniform(_settings["page_delay_min"], _settings["page_delay_max"])
                    for _ in range(int(delay)):
                        if job_state["stop_requested"]:
                            break
                        time.sleep(1)

            if job_state["stop_requested"]:
                break

            # 해시태그/위치 검색
            try:
                if is_location_search and location_pk:
                    data = {"tab": "recent"}
                    if cursor:
                        data["page"] = str(total_pages - 1)
                        data["max_id"] = cursor
                    result = client.private_request(f"locations/{location_pk}/sections/", data=data)
                else:
                    data = {"tab": tab}
                    if cursor:
                        data["page"] = str(total_pages - 1)
                        data["max_id"] = cursor
                    result = client.private_request(f"tags/{tag}/sections/", data=data)
            except Exception as e:
                err_msg = str(e)[:200]
                log.warning(f"[self-collect:{job_id}] p{total_pages} @{acc_uname} 에러: {err_msg}")

                err_lower = err_msg.lower()
                is_challenge = any(k in err_lower for k in ["challenge", "checkpoint"])
                is_login_expired = "login_required" in err_lower
                is_block = any(k in err_lower for k in ["login_required", "challenge", "feedback_required", "consent_required", "checkpoint"])
                # 캡차/셀피 → 자동 해제 절대 불가
                is_unsolvable = any(k in err_lower for k in ["recaptcha", "captcha", "selfie"])

                # 챌린지/로그인만료 → 세션 삭제 후 TOTP 재로그인으로 자동 해제 시도
                # (feedback_required, consent_required, captcha 등은 재로그인으로 해결 불가)
                if (is_challenge or is_login_expired) and not is_unsolvable:
                    accounts_now = _load_self_accounts()
                    acc_data = next((a for a in accounts_now if a["username"] == acc_uname), None)
                    totp_secret = acc_data.get("totp_secret", "") if acc_data else ""

                    if totp_secret:
                        _update(status=f"@{acc_uname} 챌린지 — UUID 보존 TOTP 재로그인 시도...")
                        log.info(f"[self-collect:{job_id}] @{acc_uname} 챌린지 → UUID 보존 재로그인")
                        try:
                            # ★ 세션 삭제 금지! UUID 보존한 채 재로그인
                            from instagrapi import Client as _Client
                            new_cl = _Client()
                            new_cl.delay_range = [_settings["login_delay_min"], _settings["login_delay_max"]]
                            _login_with_session(new_cl, acc_uname, acc_data["password"], totp_secret, proxy)
                            new_cl.delay_range = [1, 3]  # 크롤링용 딜레이
                            # 재로그인 성공 → 기존 클라이언트 교체, 수집 계속
                            client = new_cl
                            _update_account_status(acc_uname, "active", "")
                            _update(status=f"@{acc_uname} TOTP 재로그인 성공 — 수집 재개")
                            log.info(f"[self-collect:{job_id}] @{acc_uname} TOTP 재로그인 성공!")
                            time.sleep(5)
                            continue
                        except Exception as e2:
                            resolve_err = str(e2)[:150]
                            log.warning(f"[self-collect:{job_id}] @{acc_uname} TOTP 재로그인 실패: {resolve_err}")
                            _update_account_status(acc_uname, "blocked",
                                f"자동 해제 실패 — {_translate_insta_error(resolve_err)}")
                    else:
                        _update_account_status(acc_uname, "blocked",
                            f"TOTP 없음 — {_translate_insta_error(err_msg[:100])}")

                elif is_block:
                    _update_account_status(acc_uname, "blocked", _translate_insta_error(err_msg[:100]))

                if is_block:
                    _update(status=f"@{acc_uname} 차단 — 다음 계정으로 전환")
                    _release_current()
                    account_pages = 0

                    client, acc_uname = _login_next()
                    if not client:
                        _update(done=True, status=f"사용 가능한 계정 없음 — 신규 {new_count}명 / {total_posts}게시물",
                                new=new_count, dup=dup_count, posts=total_posts, pages=total_pages)
                        break
                    continue
                else:
                    time.sleep(10)
                    continue

            next_max_id = result.get("next_max_id")
            sections = result.get("sections", [])

            # 연관 해시태그 추출 (첫 페이지에서)
            if total_pages == 1 and not is_location_search:
                related = []
                for rt in result.get("related_tags", []):
                    if isinstance(rt, str):
                        related.append({"name": rt, "count": 0})
                    elif isinstance(rt, dict):
                        rname = rt.get("name", "")
                        rcount = rt.get("media_count", 0) or rt.get("count", 0) or 0
                        if rname and rname != tag:
                            related.append({"name": rname, "count": rcount})
                if related:
                    _update(related_tags=related[:20])

            # 게시물에서 유저 추출 + 게시물 저장
            batch_users = []
            batch_posts = []
            page_new = 0
            for s in sections:
                for m in s.get("layout_content", {}).get("medias", []):
                    total_posts += 1
                    media = m.get("media", {})
                    u = media.get("user", {})
                    user_pk = str(u.get("pk", ""))
                    username = u.get("username", "")

                    # 게시물 데이터 저장용
                    media_id = str(media.get("pk", "") or media.get("id", ""))
                    caption_obj = media.get("caption") or {}
                    caption_text = caption_obj.get("text", "") if isinstance(caption_obj, dict) else ""
                    image_versions = media.get("image_versions2", {}).get("candidates", [])
                    thumbnail_url = image_versions[0].get("url", "") if image_versions else ""
                    carousel = media.get("carousel_media", [])
                    if not thumbnail_url and carousel:
                        first_img = carousel[0].get("image_versions2", {}).get("candidates", [])
                        thumbnail_url = first_img[0].get("url", "") if first_img else ""

                    if media_id:
                        batch_posts.append({
                            "media_id": media_id,
                            "username": username,
                            "full_name": u.get("full_name", "") or "",
                            "profile_pic_url": u.get("profile_pic_url", "") or "",
                            "thumbnail_url": thumbnail_url,
                            "caption": caption_text[:500],
                            "like_count": media.get("like_count", 0) or 0,
                            "comment_count": media.get("comment_count", 0) or 0,
                            "taken_at": media.get("taken_at", 0) or 0,
                            "media_type": media.get("media_type", 1),
                            "carousel_count": len(carousel) if carousel else 0,
                            "job_id": job_id,
                            "hashtag": tag,
                        })

                    if not user_pk or not username or user_pk in seen_user_pks:
                        continue
                    seen_user_pks.add(user_pk)

                    if user_pk in banned_pks:
                        continue

                    full_name = u.get("full_name", "") or ""

                    # 블랙리스트 필터 → 밴 테이블에 등록 (영구 제외)
                    bl_reason = _is_blacklisted(username, full_name, _blacklist)
                    if bl_reason:
                        skip_count += 1
                        try:
                            # DB에 밴 등록 (먼저 upsert로 레코드 생성 후 밴)
                            upsert_influencer({
                                "pk": user_pk,
                                "username": username,
                                "full_name": full_name,
                                "profile_pic_url": u.get("profile_pic_url", "") or "",
                                "source_hashtag": tag,
                            })
                            ban_influencer(user_pk, f"블랙리스트: {bl_reason}")
                            banned_pks.add(user_pk)
                        except Exception:
                            pass
                        continue

                    is_new = user_pk not in existing_pks

                    if is_new:
                        new_count += 1
                        page_new += 1
                        try:
                            upsert_influencer({
                                "pk": user_pk,
                                "username": username,
                                "full_name": full_name,
                                "profile_pic_url": u.get("profile_pic_url", "") or "",
                                "source_hashtag": tag,
                            })
                            existing_pks.add(user_pk)
                        except Exception as db_err:
                            log.warning(f"[self-collect:{job_id}] DB upsert 실패: {db_err}")
                    else:
                        dup_count += 1

                    batch_users.append({"username": username, "full_name": full_name, "is_new": is_new})

            # 게시물 파일 저장
            if batch_posts:
                try:
                    _save_self_posts(job_id, tag, batch_posts)
                except Exception:
                    pass

            cursor = next_max_id
            account_pages += 1

            # 수집된 유저를 히스토리에 저장
            if batch_users:
                try:
                    _jobs = _load_self_jobs()
                    for _j in _jobs:
                        if _j["id"] == job_id:
                            if "users" not in _j:
                                _j["users"] = []
                            _j["users"].extend(batch_users)
                            break
                    _save_self_jobs(_jobs)
                except Exception:
                    pass

            _update(
                new=new_count, dup=dup_count, posts=total_posts, pages=total_pages,
                status=f"p{total_pages} @{acc_uname} ({account_pages}/{MAX_PAGES_PER_ACCOUNT}) — 신규 {new_count}명 / 중복 {dup_count}명 / 제외 {skip_count}명 / {total_posts}게시물 (+{page_new})",
                current_account=acc_uname,
                users=batch_users,
            )

            # 선제적 계정 교체
            if account_pages >= MAX_PAGES_PER_ACCOUNT and cursor:
                log.info(f"[self-collect:{job_id}] @{acc_uname} {account_pages}페이지 도달 — 선제 교체")
                _update_account_status(acc_uname, "resting")
                try:
                    client.logout()
                except Exception:
                    pass
                _release_current()
                account_pages = 0

                # 쿨다운: 1초 간격으로 중지 확인
                _update(status=f"계정 교체 중... {_settings['account_switch_delay']}초 대기")
                for _ in range(int(_settings["account_switch_delay"])):
                    if job_state["stop_requested"]:
                        break
                    time.sleep(1)

                client, acc_uname = _login_next()
                if not client:
                    _update(done=True, status=f"사용 가능한 계정 없음 — 신규 {new_count}명",
                            new=new_count, dup=dup_count, posts=total_posts, pages=total_pages)
                    break
                _update(status=f"@{acc_uname}로 교체 — 커서 이어받기")

            if new_count >= target_users:
                _update(done=True, status=f"목표 달성! 신규 {new_count}명 / 제외 {skip_count}명 / {total_posts}게시물 / {total_pages}페이지",
                        new=new_count, dup=dup_count, posts=total_posts, pages=total_pages)
                break

            if not cursor:
                _update(done=True, status=f"마지막 페이지 도달 — 신규 {new_count}명 / 제외 {skip_count}명 / {total_posts}게시물",
                        new=new_count, dup=dup_count, posts=total_posts, pages=total_pages)
                break

    except Exception as e:
        log.error(f"[self-collect:{job_id}] worker 에러: {e}")
        _update(done=True, status=f"에러: {str(e)[:150]}", error=str(e)[:150])

    if job_state["stop_requested"] and not progress.get("done"):
        _update(done=True, stopped=True, status=f"사용자 중지 — 신규 {new_count}명 / 제외 {skip_count}명 / {total_posts}게시물 / {total_pages}페이지",
                new=new_count, dup=dup_count, posts=total_posts, pages=total_pages)

    # 정리
    _release_current()
    _finish_job(new_count, dup_count, total_posts, total_pages,
                progress.get("error"))
    try:
        if client:
            client.logout()
    except Exception:
        pass

    # 잠시 후 running dict에서 제거 (SSE가 done 읽을 시간)
    time.sleep(5)
    with _self_lock:
        _self_jobs_running.pop(job_id, None)


# ═══════════════════════════════════════════════════════
# 광고주 계정 관리 (관리자)
# ═══════════════════════════════════════════════════════

@app.get("/advertisers", response_class=HTMLResponse)
def advertisers_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    advs = get_advertisers()
    return templates.TemplateResponse("advertiser_list.html", {
        "request": request, "user": user, "advertisers": advs,
        "plans": PLANS, "now": time.time(),
    })

@app.post("/advertisers/add")
def add_advertiser_route(
    username: str = Form(...), password: str = Form(...),
    company_name: str = Form(default=""),
    hashtag_access: str = Form(default=""),
    min_followers: int = Form(default=0),
    only_approved: int = Form(default=1),
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    try:
        db_add_advertiser(username.strip(), pw_hash, company_name, hashtag_access, min_followers, only_approved)
    except Exception as e:
        log.error(f"광고주 추가 실패: {e}")
    return RedirectResponse("/advertisers?added=1", 302)

@app.post("/advertisers/delete")
def delete_advertiser_route(adv_id: int = Form(...), session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    db_delete_advertiser(adv_id)
    return RedirectResponse("/advertisers", 302)

@app.post("/advertisers/{adv_id}/plan")
def set_advertiser_plan(
    adv_id: int,
    plan: str = Form(...),
    months: int = Form(default=1),
    hashtag_access: str = Form(default=""),
    min_followers: int = Form(default=0),
    only_approved: int = Form(default=1),
    session_id: Optional[str] = Cookie(default=None)
):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    plan_info = PLANS.get(plan, PLANS["free"])
    expires = time.time() + months * 30 * 86400
    update_advertiser_plan(
        adv_id, plan, expires,
        plan_info["per_hashtag"],
        hashtag_access or None,
        min_followers or None,
        only_approved
    )
    return RedirectResponse("/advertisers?updated=1", 302)


# ═══════════════════════════════════════════════════════
# 광고주 로그인 / 대시보드
# ═══════════════════════════════════════════════════════

@app.get("/advertiser/login", response_class=HTMLResponse)
def adv_login_page(request: Request):
    return templates.TemplateResponse("advertiser_login.html", {"request": request})

@app.post("/advertiser/login")
def adv_login(request: Request, username: str = Form(...), password: str = Form(...)):
    ip = _get_client_ip(request)
    if not _check_login_allowed(ip):
        log.warning(f"광고주 로그인 brute-force 차단: {ip}")
        return templates.TemplateResponse("advertiser_login.html", {
            "request": request, "error": "너무 많은 로그인 시도입니다. 15분 후 다시 시도해주세요."
        }, status_code=429)
    row = get_advertiser_by_username(username.strip())
    if row and bcrypt.checkpw(password.encode(), (row.get("password_hash") or "").encode()):
        _login_attempts.pop(ip, None)
        token = _make_jwt({"role": "advertiser", "username": username, "adv_id": row.get("id")})
        res = RedirectResponse("/advertiser", status_code=302)
        res.set_cookie("adv_session_id", token, httponly=True, max_age=604800,
                        samesite="lax", secure=_IS_PROD)
        return res
    _record_login_attempt(ip)
    remaining = _LOGIN_MAX_ATTEMPTS - len(_login_attempts.get(ip, []))
    return templates.TemplateResponse("advertiser_login.html", {
        "request": request, "error": f"아이디 또는 비밀번호가 올바르지 않습니다. (남은 시도: {remaining}회)"
    }, status_code=400)

@app.get("/advertiser/logout")
def adv_logout(adv_session_id: Optional[str] = Cookie(default=None)):
    adv_sessions.pop(adv_session_id, None)
    res = RedirectResponse("/advertiser/login", 302)
    res.delete_cookie("adv_session_id")
    return res

# ─── 찜 API ───────────────────────────────────────────────────────────
@app.post("/advertiser/favorites/toggle")
def adv_toggle_favorite(
    influencer_pk: str = Form(...),
    adv_session_id: Optional[str] = Cookie(default=None)
):
    adv = get_adv_user(adv_session_id)
    if not adv: return JSONResponse({"error": "로그인 필요"}, 401)
    added = toggle_favorite(adv["id"], influencer_pk)
    return JSONResponse({"ok": True, "favorited": added})

@app.get("/advertiser/favorites", response_class=HTMLResponse)
def adv_favorites_page(request: Request, adv_session_id: Optional[str] = Cookie(default=None)):
    adv = get_adv_user(adv_session_id)
    if not adv: return RedirectResponse("/advertiser/login", 302)
    favs = get_favorites(adv["id"])
    pks = [f["influencer_pk"] for f in favs]
    rows = [get_influencer(pk) for pk in pks if pk]
    rows = [r for r in rows if r]
    campaigns = get_campaigns(adv["id"])
    return templates.TemplateResponse("advertiser_favorites.html", {
        "request": request, "adv": adv, "rows": rows,
        "fav_pks": set(pks), "campaigns": campaigns,
    })


# ─── 캠페인 API ────────────────────────────────────────────────────────
@app.post("/advertiser/campaigns/create")
def adv_create_campaign(
    name: str = Form(...),
    description: str = Form(default=""),
    budget: int = Form(default=0),
    adv_session_id: Optional[str] = Cookie(default=None)
):
    adv = get_adv_user(adv_session_id)
    if not adv: return JSONResponse({"error": "로그인 필요"}, 401)
    camp_id = create_campaign(adv["id"], name, description, budget)
    return JSONResponse({"ok": True, "id": camp_id})

@app.get("/advertiser/campaigns", response_class=HTMLResponse)
def adv_campaigns_page(request: Request, adv_session_id: Optional[str] = Cookie(default=None)):
    adv = get_adv_user(adv_session_id)
    if not adv: return RedirectResponse("/advertiser/login", 302)
    camps = get_campaigns(adv["id"])
    return templates.TemplateResponse("advertiser_campaigns.html", {
        "request": request, "adv": adv, "campaigns": camps,
    })

@app.get("/advertiser/campaigns/{camp_id}", response_class=HTMLResponse)
def adv_campaign_detail(camp_id: int, request: Request, adv_session_id: Optional[str] = Cookie(default=None)):
    adv = get_adv_user(adv_session_id)
    if not adv: return RedirectResponse("/advertiser/login", 302)
    camp = get_campaign(camp_id)
    if not camp or camp.get("advertiser_id") != adv["id"]: raise HTTPException(403)
    cinfs = get_campaign_influencers(camp_id)
    # 인플루언서 상세 정보 조인
    rows = []
    total_price = 0
    for ci in cinfs:
        inf = get_influencer(ci["influencer_pk"])
        if inf:
            manual = get_manual(ci["influencer_pk"])
            rows.append({**ci, "inf": inf, "manual": manual})
            total_price += ci.get("price", 0) or 0
    return templates.TemplateResponse("advertiser_campaign_detail.html", {
        "request": request, "adv": adv, "camp": camp,
        "rows": rows, "total_price": total_price,
    })

@app.post("/advertiser/campaigns/{camp_id}/add")
def adv_add_to_campaign(
    camp_id: int,
    influencer_pk: str = Form(...),
    content_type: str = Form(default="feed"),
    price: int = Form(default=0),
    note: str = Form(default=""),
    adv_session_id: Optional[str] = Cookie(default=None)
):
    adv = get_adv_user(adv_session_id)
    if not adv: return JSONResponse({"error": "로그인 필요"}, 401)
    camp = get_campaign(camp_id)
    if not camp or camp.get("advertiser_id") != adv["id"]: return JSONResponse({"error": "권한 없음"}, 403)
    added = add_to_campaign(camp_id, influencer_pk, content_type, price, note)
    return JSONResponse({"ok": True, "added": added})

@app.post("/advertiser/campaigns/{camp_id}/remove")
def adv_remove_from_campaign(
    camp_id: int,
    influencer_pk: str = Form(...),
    adv_session_id: Optional[str] = Cookie(default=None)
):
    adv = get_adv_user(adv_session_id)
    if not adv: return JSONResponse({"error": "로그인 필요"}, 401)
    remove_from_campaign(camp_id, influencer_pk)
    return JSONResponse({"ok": True})

@app.post("/advertiser/campaigns/{camp_id}/delete")
def adv_delete_campaign(camp_id: int, adv_session_id: Optional[str] = Cookie(default=None)):
    adv = get_adv_user(adv_session_id)
    if not adv: return JSONResponse({"error": "로그인 필요"}, 401)
    delete_campaign(camp_id)
    return RedirectResponse("/advertiser/campaigns", 302)


@app.get("/advertiser", response_class=HTMLResponse)
def adv_dashboard(
    request: Request,
    q: str = "",
    hashtag: str = "",
    min_f: Optional[str] = None,
    max_f: Optional[str] = None,
    main_category: str = "",
    can_live: int = 0,
    sort: str = "follower_count",
    order: str = "desc",
    page: int = Query(1, gt=0),
    per_page: int = Query(30, gt=0),
    adv_session_id: Optional[str] = Cookie(default=None)
):
    adv = get_adv_user(adv_session_id)
    if not adv: return RedirectResponse("/advertiser/login", 302)

    _min_f = int(min_f) if min_f and min_f.isdigit() else None
    _max_f = int(max_f) if max_f and max_f.isdigit() else None

    # 광고주 접근 해시태그 필터링
    allowed_hashtags = adv.get("hashtag_access", "")
    effective_hashtag = hashtag
    if allowed_hashtags and not hashtag:
        effective_hashtag = allowed_hashtags.split(",")[0].strip()

    only_approved = bool(adv.get("only_approved", 1))
    min_f_eff = max(_min_f or 0, adv.get("min_followers", 0)) or None

    total, rows = get_influencers(
        keyword=q, hashtag_filter=effective_hashtag,
        min_f=min_f_eff, max_f=_max_f,
        only_approved=only_approved,
        main_category=main_category, can_live=bool(can_live),
        sort=sort, order=order, page=page, per_page=per_page
    )
    total_pages = max(1, (total + per_page - 1) // per_page)
    allowed_list = [h.strip() for h in allowed_hashtags.split(",") if h.strip()] if allowed_hashtags else []
    adv_id = adv.get("id", 0)
    fav_pks = get_favorite_pks(adv_id)
    campaigns = get_campaigns(adv_id)

    return templates.TemplateResponse("advertiser_dashboard.html", {
        "request": request, "adv": adv,
        "rows": rows, "total": total, "total_pages": total_pages,
        "page": page, "per_page": per_page,
        "q": q, "hashtag": effective_hashtag,
        "min_f": min_f, "max_f": max_f,
        "main_category": main_category, "can_live": can_live,
        "sort": sort, "order": order,
        "allowed_hashtags": allowed_list,
        "fav_pks": fav_pks,
        "campaigns": campaigns,
    })

@app.get("/advertiser/influencers/{pk}", response_class=HTMLResponse)
def adv_influencer_detail(pk: str, request: Request,
                           adv_session_id: Optional[str] = Cookie(default=None)):
    adv = get_adv_user(adv_session_id)
    if not adv: return RedirectResponse("/advertiser/login", 302)
    inf = get_influencer(pk)
    if not inf: raise HTTPException(404)
    manual = get_manual(pk)
    # 광고주에게는 연락처 숨김 (is_approved만 공개)
    manual_safe = {
        "can_live": manual.get("can_live"),
        "live_platforms": manual.get("live_platforms"),
        "feed_price": manual.get("feed_price"),
        "reel_price": manual.get("reel_price"),
        "story_price": manual.get("story_price"),
        "main_category": manual.get("main_category"),
        "sub_categories": manual.get("sub_categories"),
        "target_gender": manual.get("target_gender"),
        "target_age": manual.get("target_age"),
        "target_region": manual.get("target_region"),
        "collab_types": manual.get("collab_types"),
        "past_brands": manual.get("past_brands"),
        "quality_score": manual.get("quality_score"),
    }
    for key in ["top_posts_likes", "top_posts_comments", "top_reels_views"]:
        try: inf[key] = json.loads(inf.get(key) or "[]")
        except: inf[key] = []
    try:
        raw = json.loads(inf.get("top_hashtags") or "[]")
        inf["top_hashtags"] = [{"tag": t, "count": 1} for t in raw] if raw and isinstance(raw[0], str) else raw
    except:
        inf["top_hashtags"] = []
    posts = get_influencer_posts(pk)
    recent_reels = get_influencer_reels(pk, sort="recent", limit=12)
    popular_reels = get_influencer_reels(pk, sort="popular", limit=12)
    # 활동 유형 분석 (광고주 뷰에서도 동일하게 표시)
    activity = _analyze_activity(inf.get("top_hashtags", []))
    return templates.TemplateResponse("influencer_detail.html", {
        "request": request, "adv": adv, "user": None,
        "inf": inf, "manual": manual_safe, "posts": posts,
        "is_advertiser_view": True,
        "activity": activity,
        "recent_reels": recent_reels,
        "popular_reels": popular_reels,
    })



# ═══════════════════════════════════════════════════════
# 설정
# ═══════════════════════════════════════════════════════

@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    return templates.TemplateResponse("settings.html", {
        "request": request, "user": user,
        "insta_username": INSTA_CFG["username"],
        "extension_api_key": EXTENSION_API_KEY,
        "plans": PLANS,
    })

@app.post("/settings/save")
def settings_save(
    insta_username: str = Form(...),
    insta_password: str = Form(...),
    insta_totp: str = Form(default=""),
    admin_password: str = Form(default=""),
    session_id: Optional[str] = Cookie(default=None)
):
    global ADMIN_PW_HASH
    user = get_user(session_id)
    if not user: return RedirectResponse("/login", 302)
    INSTA_CFG["username"] = insta_username
    INSTA_CFG["password"] = insta_password
    INSTA_CFG["totp"] = insta_totp
    if admin_password:
        ADMIN_PW_HASH = bcrypt.hashpw(admin_password.encode(), bcrypt.gensalt())

    env_path = os.path.join(os.path.dirname(__file__), ".env")
    lines = open(env_path).readlines() if os.path.exists(env_path) else []
    keys = {
        "INSTA_USERNAME": insta_username, "INSTA_PASSWORD": insta_password,
        "INSTA_TOTP": insta_totp,
    }
    if admin_password:
        keys["ADMIN_PASSWORD"] = admin_password
    existing = {l.split("=")[0]: i for i, l in enumerate(lines) if "=" in l}
    for k, v in keys.items():
        if k in existing: lines[existing[k]] = f"{k}={v}\n"
        else: lines.append(f"{k}={v}\n")
    try:
        open(env_path, "w").writelines(lines)
    except OSError:
        pass  # Vercel 환경에서는 .env 쓰기 불가
    return RedirectResponse("/settings?saved=1", 302)

@app.post("/settings/test-account")
async def test_insta_account(session_id: Optional[str] = Cookie(default=None)):
    user = get_user(session_id)
    if not user: raise HTTPException(403)
    try:
        import pyotp
        totp_code = ""
        if INSTA_CFG.get("totp"):
            totp_code = pyotp.TOTP(INSTA_CFG["totp"]).now()
        return JSONResponse({"ok": True, "totp_code": totp_code, "username": INSTA_CFG["username"]})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
